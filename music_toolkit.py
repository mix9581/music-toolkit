#!/usr/bin/env python3
"""
Music Toolkit — 音乐搜索/下载/飞书推送工具包。

封装 go-music-dl HTTP API，提供搜索、歌词、下载、飞书卡片推送等功能。
设计理念: 单文件 + CLI + 可 import，配合 SKILL.md 让 AI 一看就会用。

用法 - 作为模块:
    from music_toolkit import MusicClient
    client = MusicClient()
    songs = client.search_songs("晴天")

用法 - 作为 CLI:
    python music_toolkit.py search "晴天"
    python music_toolkit.py search "晴天" --source qq --source netease
    python music_toolkit.py lyrics 0042rlGx2WHBrG qq
    python music_toolkit.py detail 0042rlGx2WHBrG qq
    python music_toolkit.py download 0042rlGx2WHBrG qq
    python music_toolkit.py platforms

依赖: requests (pip install requests)
"""

import os
import re
import sys
import json
import html
import argparse
from pathlib import Path
from dataclasses import dataclass, field, asdict
from typing import Optional

try:
    import requests
except ImportError:
    print("Error: 'requests' package required. Run: pip install requests", file=sys.stderr)
    sys.exit(1)

__version__ = "0.1.0"

# ─── Constants ────────────────────────────────────────────────────────────────

DEFAULT_BASE_URL = "http://localhost:8080"
DEFAULT_KUGOU_API_URL = "http://localhost:3000"  # KuGouMusicApi 服务地址
DEFAULT_DOWNLOAD_DIR = "./downloads"
DEFAULT_COOKIE_FILE = ".music_cookies.json"

PLATFORMS = {
    "netease": "网易云音乐",
    "qq": "QQ音乐",
    "kugou": "酷狗音乐",
    "kuwo": "酷我音乐",
    "migu": "咪咕音乐",
    "qianqian": "千千音乐",
    "soda": "Soda音乐",
    "fivesing": "5sing",
    "jamendo": "Jamendo (CC)",
    "joox": "JOOX",
    "bilibili": "哔哩哔哩",
}

ALL_SOURCES = list(PLATFORMS.keys())

# Cookie 环境变量映射
COOKIE_ENV_VARS = {
    "netease": "NETEASE_COOKIE",
    "qq": "QQ_MUSIC_COOKIE",
    "kugou": "KUGOU_COOKIE",
    "kuwo": "KUWO_COOKIE",
    "migu": "MIGU_COOKIE",
    "bilibili": "BILIBILI_COOKIE",
}


# ─── Data Models ──────────────────────────────────────────────────────────────

@dataclass(frozen=True)
class Song:
    """歌曲数据模型"""
    id: str
    source: str
    name: str
    artist: str
    duration: int
    cover: str
    album: str = ""
    lyrics: str = ""
    url: str = ""
    size: str = ""
    bitrate: str = ""
    link: str = ""
    score: float = 0.0
    extra: dict = field(default_factory=dict)

    @property
    def duration_str(self) -> str:
        """格式化时长 (e.g. 4:29)"""
        if self.duration <= 0:
            return "0:00"
        minutes = self.duration // 60
        seconds = self.duration % 60
        return f"{minutes}:{seconds:02d}"

    @property
    def source_name(self) -> str:
        """平台中文名"""
        return PLATFORMS.get(self.source, self.source)

    def to_dict(self) -> dict:
        """转为字典（含计算属性）"""
        result = asdict(self)
        result["duration_str"] = self.duration_str
        result["source_name"] = self.source_name
        return result


@dataclass(frozen=True)
class Playlist:
    """歌单数据模型"""
    id: str
    source: str
    name: str
    cover: str
    track_count: int = 0
    play_count: int = 0
    creator: str = ""
    description: str = ""

    @property
    def source_name(self) -> str:
        return PLATFORMS.get(self.source, self.source)

    def to_dict(self) -> dict:
        result = asdict(self)
        result["source_name"] = self.source_name
        return result


@dataclass(frozen=True)
class InspectResult:
    """歌曲资源检查结果"""
    valid: bool
    url: str
    size: str
    bitrate: str

    def to_dict(self) -> dict:
        return asdict(self)


@dataclass(frozen=True)
class SongDetailInfo:
    """从平台页面/API 抓取的歌曲详情（含收藏、评论、分享等统计）。"""
    song_id: str
    platform: str
    name: str
    artist: str
    duration: int          # 时长（秒）
    cover: str             # 封面图 URL
    album: str
    album_id: str          # 专辑 ID
    publish_date: str      # 发布日期 YYYY-MM-DD
    favorites: int         # 收藏数
    comments: int          # 评论数
    shares: int            # 分享数
    plays: int             # 播放/使用数
    audio_url: str         # 音频 CDN 直链（临时有效，通常 24h）
    lyrics_lrc: str        # LRC 格式歌词（含时间轴）
    genre: str             # 曲风
    language: str          # 语言代码（如 ZH / EN）
    composers: str         # 作曲（逗号分隔）
    lyricists: str         # 作词（逗号分隔）
    qualities: str         # 音质选项（如 "medium(68k) / higher(132k) / lossless"）
    share_url: str         # 原始分享链接
    resolved_url: str      # 解析后的完整 URL
    extra: dict = field(default_factory=dict)  # 平台原始数据（含更多字段）

    @property
    def duration_str(self) -> str:
        if self.duration <= 0:
            return "0:00"
        minutes = self.duration // 60
        seconds = self.duration % 60
        return f"{minutes}:{seconds:02d}"

    @property
    def lyrics_text(self) -> str:
        """纯文本歌词（去除时间轴）"""
        lines = []
        for line in self.lyrics_lrc.splitlines():
            text = re.sub(r'\[\d+:\d+[\.\d]*\]', '', line).strip()
            if text:
                lines.append(text)
        return "\n".join(lines)

    def to_dict(self) -> dict:
        result = asdict(self)
        result["duration_str"] = self.duration_str
        result["lyrics_text"] = self.lyrics_text
        return result


@dataclass(frozen=True)
class PlaylistDetailInfo:
    """歌单详情数据模型（数据监控用，不含歌曲下载）"""
    playlist_id: str
    platform: str
    title: str
    creator: str
    cover: str
    track_count: int
    create_time: str      # YYYY-MM-DD
    update_time: str      # YYYY-MM-DD
    description: str
    tracks: tuple         # tuple[SongDetailInfo]
    share_url: str
    resolved_url: str
    extra: dict = field(default_factory=dict)

    @property
    def source_name(self) -> str:
        return PLATFORMS.get(self.platform, self.platform)

    def to_dict(self) -> dict:
        return {
            "playlist_id": self.playlist_id,
            "platform": self.platform,
            "source_name": self.source_name,
            "title": self.title,
            "creator": self.creator,
            "cover": self.cover,
            "track_count": self.track_count,
            "create_time": self.create_time,
            "update_time": self.update_time,
            "description": self.description,
            "share_url": self.share_url,
            "resolved_url": self.resolved_url,
            "extra": self.extra,
            "tracks": [t.to_dict() for t in self.tracks],
        }


@dataclass
class DownloadResult:
    """单首歌曲下载结果"""
    song: Song
    success: bool
    filepath: Optional[Path] = None
    lrc_path: Optional[Path] = None
    txt_path: Optional[Path] = None
    actual_source: str = ""
    error: str = ""

    def to_dict(self) -> dict:
        d = {
            "name": self.song.name,
            "artist": self.song.artist,
            "source": self.song.source,
            "actual_source": self.actual_source or self.song.source,
            "success": self.success,
        }
        if self.filepath:
            d["filepath"] = str(self.filepath)
        if self.lrc_path:
            d["lrc_path"] = str(self.lrc_path)
        if self.txt_path:
            d["txt_path"] = str(self.txt_path)
        if self.error:
            d["error"] = self.error
        return d


# ─── HTML Parsing Helpers ─────────────────────────────────────────────────────

def _parse_song_cards(html_text: str) -> list[Song]:
    """从 go-music-dl 搜索结果 HTML 中解析歌曲列表。

    HTML 结构:
      <li class="song-card" data-id="xxx" data-source="qq"
          data-duration="278" data-name="晴天" data-artist="周杰伦"
          data-cover="https://..." data-extra='{"songmid":"xxx"}'>
    """
    songs = []
    # Match each song-card <li> block
    card_pattern = re.compile(
        r'<li\s+class="song-card"[^>]*?'
        r'data-id="([^"]*)"[^>]*?'
        r'data-source="([^"]*)"[^>]*?'
        r'data-duration="([^"]*)"[^>]*?'
        r'data-name="([^"]*)"[^>]*?'
        r'data-artist="([^"]*)"[^>]*?'
        r'data-cover="([^"]*)"',
        re.DOTALL,
    )
    extra_pattern = re.compile(r"data-extra='([^']*)'")
    # 专辑名在 artist-line 里: <i ...></i> 歌手名 &nbsp;•&nbsp; 专辑名
    album_pattern = re.compile(
        r'<div\s+class="artist-line">[^<]*<i[^>]*></i>\s*[^&]*&nbsp;•&nbsp;\s*([^<\n]+)',
    )

    for match in card_pattern.finditer(html_text):
        song_id, source, duration_str, name, artist, cover = match.groups()

        # Grab a wider block to extract extra + album
        block_end = min(match.start() + 2000, len(html_text))
        card_html = html_text[match.start():block_end]

        # Try to extract extra JSON data
        extra = {}
        extra_match = extra_pattern.search(card_html)
        if extra_match:
            try:
                extra_raw = html.unescape(extra_match.group(1))
                extra = json.loads(extra_raw)
            except (json.JSONDecodeError, ValueError):
                pass

        # Try to extract album name
        album = ""
        album_match = album_pattern.search(card_html)
        if album_match:
            album = html.unescape(album_match.group(1).strip())

        songs.append(Song(
            id=html.unescape(song_id),
            source=html.unescape(source),
            name=html.unescape(name),
            artist=html.unescape(artist),
            duration=int(duration_str) if duration_str.isdigit() else 0,
            cover=html.unescape(cover),
            album=album,
            extra=extra,
        ))

    return songs


def _parse_playlist_cards(html_text: str) -> list[Playlist]:
    """从 go-music-dl 歌单搜索结果 HTML 中解析歌单列表。

    HTML 结构:
      <div class="playlist-card" onclick="location.href='/music/playlist?id=xxx&source=netease'">
        <div class="playlist-cover"><img src="..." ...></div>
        <div class="playlist-meta">
          <div class="playlist-title">名称</div>
          <div class="playlist-author"><i ...></i> 创建者</div>
          <div class="playlist-count">共 147 首</div>
        </div>
      </div>
    """
    playlists = []

    # Extract each playlist-card block
    card_pattern = re.compile(
        r'<div\s+class="playlist-card"[^>]*onclick="location\.href=\'[^\']*\?id=([^&]*)&(?:amp;)?source=([^\']*)\'"[^>]*>',
        re.DOTALL,
    )
    img_pattern = re.compile(r'<img\s+src="([^"]*)"')
    title_pattern = re.compile(r'<div\s+class="playlist-title">\s*(.*?)(?:\s*<a\s)', re.DOTALL)
    author_pattern = re.compile(r'<div\s+class="playlist-author">[^<]*<i[^>]*></i>\s*([^<]*)</div>')
    count_pattern = re.compile(r'<div\s+class="playlist-count">共\s*(\d+)\s*首</div>')

    # Split HTML by playlist-card blocks
    card_splits = re.split(r'(?=<div\s+class="playlist-card")', html_text)

    for block in card_splits:
        card_match = card_pattern.search(block)
        if not card_match:
            continue

        playlist_id = card_match.group(1)
        source = card_match.group(2)

        img_match = img_pattern.search(block)
        cover = img_match.group(1) if img_match else ""

        title_match = title_pattern.search(block)
        name = title_match.group(1).strip() if title_match else ""
        # Clean HTML entities and whitespace
        name = html.unescape(re.sub(r'\s+', ' ', name).strip())

        author_match = author_pattern.search(block)
        creator = author_match.group(1).strip() if author_match else ""

        count_match = count_pattern.search(block)
        track_count = int(count_match.group(1)) if count_match else 0

        playlists.append(Playlist(
            id=playlist_id,
            source=source,
            name=name,
            cover=cover,
            track_count=track_count,
            creator=creator,
        ))

    return playlists


# ─── MusicClient: go-music-dl HTTP Client ─────────────────────────────────────

class MusicClientError(Exception):
    """MusicClient 业务异常"""
    pass


class MusicClient:
    """go-music-dl HTTP API 客户端。

    Args:
        base_url: go-music-dl 服务地址 (默认 http://localhost:8080)
        timeout: HTTP 请求超时秒数
    """

    def __init__(
        self,
        base_url: str = None,
        timeout: int = 30,
    ):
        raw_url = base_url or os.environ.get("GO_MUSIC_DL_URL", DEFAULT_BASE_URL)
        self.base_url = raw_url.rstrip("/")
        self.timeout = timeout
        self._session = requests.Session()

    # ── Search ─────────────────────────────────────────────────────────────

    def search_songs(
        self,
        keyword: str,
        sources: list[str] = None,
    ) -> list[Song]:
        """搜索歌曲。

        Args:
            keyword: 搜索关键词
            sources: 搜索平台列表 (默认全部)

        Returns:
            歌曲列表
        """
        params = {"q": keyword, "type": "song"}
        if sources:
            params["sources"] = ",".join(sources)
        resp = self._get("/music/search", params=params)
        return _parse_song_cards(resp.text)

    def search_playlists(
        self,
        keyword: str,
        sources: list[str] = None,
    ) -> list[Playlist]:
        """搜索歌单。

        Args:
            keyword: 搜索关键词
            sources: 搜索平台列表 (默认全部)

        Returns:
            歌单列表
        """
        params = {"q": keyword, "type": "playlist"}
        if sources:
            params["sources"] = ",".join(sources)
        resp = self._get("/music/search", params=params)
        return _parse_playlist_cards(resp.text)

    # ── Song Details ───────────────────────────────────────────────────────

    def get_lyrics(self, song_id: str, source: str) -> str:
        """获取歌词 (LRC 格式文本)。

        Args:
            song_id: 歌曲 ID
            source: 音源平台

        Returns:
            LRC 格式歌词文本
        """
        resp = self._get("/music/lyric", params={"id": song_id, "source": source})
        return resp.text.strip()

    def inspect(
        self,
        song_id: str,
        source: str,
        duration: int = 0,
    ) -> InspectResult:
        """检查歌曲资源详情 (URL/大小/码率)。

        Args:
            song_id: 歌曲 ID
            source: 音源平台
            duration: 歌曲时长（秒），用于更精确的匹配

        Returns:
            InspectResult 包含 valid, url, size, bitrate
        """
        params = {"id": song_id, "source": source}
        if duration > 0:
            params["duration"] = str(duration)
        resp = self._get("/music/inspect", params=params)
        data = resp.json()
        return InspectResult(
            valid=data.get("valid", False),
            url=data.get("url", ""),
            size=data.get("size", ""),
            bitrate=data.get("bitrate", ""),
        )

    def switch_source(
        self,
        name: str,
        artist: str,
        source: str = "",
        duration: int = 0,
    ) -> Optional[Song]:
        """换源搜索 - 在其他平台查找相同歌曲。

        Args:
            name: 歌曲名
            artist: 歌手名
            source: 当前平台 (排除该平台)
            duration: 歌曲时长

        Returns:
            找到的 Song 或 None
        """
        params = {
            "name": name,
            "artist": artist,
        }
        if source:
            params["source"] = source
        if duration > 0:
            params["duration"] = str(duration)
        resp = self._get("/music/switch_source", params=params)
        if resp.status_code != 200:
            return None
        try:
            data = resp.json()
        except (json.JSONDecodeError, ValueError):
            return None

        return Song(
            id=data.get("id", ""),
            source=data.get("source", ""),
            name=data.get("name", ""),
            artist=data.get("artist", ""),
            duration=data.get("duration", 0),
            cover=data.get("cover", ""),
            album=data.get("album", ""),
            link=data.get("link", ""),
            score=data.get("score", 0.0),
        )

    # ── Download ───────────────────────────────────────────────────────────

    def download(
        self,
        song_id: str,
        source: str,
        name: str = "",
        artist: str = "",
        cover: str = "",
        embed: bool = False,
        extra: dict = None,
        save_dir: str = None,
    ) -> Path:
        """下载歌曲音频文件。

        Args:
            song_id: 歌曲 ID
            source: 音源平台
            name: 歌曲名 (用于文件名)
            artist: 歌手名 (用于文件名)
            cover: 封面 URL (embed=True 时嵌入)
            embed: 是否嵌入封面到音频文件
            extra: 额外参数 (JSON dict, go-music-dl 内部使用)
            save_dir: 保存目录

        Returns:
            保存的文件路径
        """
        target_dir = Path(save_dir or os.environ.get("DOWNLOAD_DIR", DEFAULT_DOWNLOAD_DIR))
        target_dir.mkdir(parents=True, exist_ok=True)

        params = {
            "id": song_id,
            "source": source,
        }
        if name:
            params["name"] = name
        if artist:
            params["artist"] = artist
        if cover:
            params["cover"] = cover
        if embed:
            params["embed"] = "true"
        if extra:
            params["extra"] = json.dumps(extra, ensure_ascii=False)

        resp = self._get("/music/download", params=params, stream=True)

        # Determine filename from Content-Disposition or construct one
        filename = _extract_filename(resp) or _build_filename(name, artist, "mp3")
        filepath = target_dir / _sanitize_filename(filename)

        with open(filepath, "wb") as f:
            for chunk in resp.iter_content(chunk_size=8192):
                if chunk:
                    f.write(chunk)

        return filepath

    def download_lyrics_file(
        self,
        song_id: str,
        source: str,
        name: str = "",
        artist: str = "",
        save_dir: str = None,
    ) -> Path:
        """下载歌词文件 (.lrc)。

        Args:
            song_id: 歌曲 ID
            source: 音源平台
            name: 歌曲名
            artist: 歌手名
            save_dir: 保存目录

        Returns:
            保存的歌词文件路径
        """
        target_dir = Path(save_dir or os.environ.get("DOWNLOAD_DIR", DEFAULT_DOWNLOAD_DIR))
        target_dir.mkdir(parents=True, exist_ok=True)

        params = {"id": song_id, "source": source}
        if name:
            params["name"] = name
        if artist:
            params["artist"] = artist

        resp = self._get("/music/download_lrc", params=params)

        filename = _extract_filename(resp) or _build_filename(name, artist, "lrc")
        filepath = target_dir / _sanitize_filename(filename)
        filepath.write_text(resp.text, encoding="utf-8")
        return filepath

    def download_cover(
        self,
        cover_url: str,
        name: str = "",
        artist: str = "",
        save_dir: str = None,
    ) -> Path:
        """下载封面图片。

        Args:
            cover_url: 封面图片 URL
            name: 歌曲名
            artist: 歌手名
            save_dir: 保存目录

        Returns:
            保存的封面文件路径
        """
        target_dir = Path(save_dir or os.environ.get("DOWNLOAD_DIR", DEFAULT_DOWNLOAD_DIR))
        target_dir.mkdir(parents=True, exist_ok=True)

        params = {"url": cover_url}
        if name:
            params["name"] = name
        if artist:
            params["artist"] = artist

        resp = self._get("/music/download_cover", params=params, stream=True)

        filename = _extract_filename(resp) or _build_filename(name, artist, "jpg")
        filepath = target_dir / _sanitize_filename(filename)

        with open(filepath, "wb") as f:
            for chunk in resp.iter_content(chunk_size=8192):
                if chunk:
                    f.write(chunk)

        return filepath

    # ── Playlist ───────────────────────────────────────────────────────────

    def get_playlist_songs(self, playlist_id: str, source: str) -> list[Song]:
        """获取歌单中的所有歌曲。

        Args:
            playlist_id: 歌单 ID
            source: 音源平台

        Returns:
            歌曲列表
        """
        resp = self._get("/music/playlist", params={"id": playlist_id, "source": source})
        return _parse_song_cards(resp.text)

    def download_playlist(
        self,
        playlist_id: str,
        source: str,
        save_dir: str = None,
        on_progress: callable = None,
    ) -> list["DownloadResult"]:
        """批量下载歌单中的所有歌曲（自动换源）。

        逻辑:
          1. 获取歌单歌曲列表
          2. 对每首歌: inspect → 可用则直接下载
          3. 不可用则 switch_source → inspect 换源结果 → 下载
          4. 同时下载歌词文件 (.lrc)

        Args:
            playlist_id: 歌单 ID
            source: 音源平台
            save_dir: 保存目录
            on_progress: 进度回调 fn(index, total, song_name, status_msg)

        Returns:
            每首歌的下载结果列表
        """
        songs = self.get_playlist_songs(playlist_id, source)
        if not songs:
            return []

        results = []
        total = len(songs)

        for idx, song in enumerate(songs, 1):
            if on_progress:
                on_progress(idx, total, song.name, "检查资源...")

            result = self._download_single_with_fallback(song, save_dir=save_dir)

            if on_progress:
                status = "✅" if result.success else f"❌ {result.error}"
                on_progress(idx, total, song.name, status)

            results.append(result)

        return results

    def _download_single_with_fallback(
        self,
        song: Song,
        save_dir: str = None,
    ) -> "DownloadResult":
        """下载单首歌曲，穷尽所有源直到成功。

        策略:
          1. 先尝试原始平台 inspect → 可用则下载
          2. 下载失败 (如 502) → 继续搜其他平台
          3. 原始平台不可用 → 搜索所有平台找匹配 → 逐个尝试
          4. 所有源穷尽仍失败才放弃

        Args:
            song: 歌曲对象
            save_dir: 保存目录

        Returns:
            DownloadResult
        """
        tried_sources = set()
        errors = []

        # ── 尝试 1: 原始平台 ──
        result = self._try_download_from_source(song, song, save_dir)
        if result.success:
            return result
        tried_sources.add(song.source)
        if result.error:
            errors.append(f"{song.source}: {result.error}")

        # ── 尝试 2: switch_source (API 推荐的最佳匹配) ──
        try:
            alt = self.switch_source(
                song.name, song.artist,
                source=song.source, duration=song.duration,
            )
            if alt and alt.id and alt.source not in tried_sources:
                alt_song = Song(
                    id=alt.id, source=alt.source,
                    name=alt.name or song.name,
                    artist=alt.artist or song.artist,
                    duration=alt.duration, cover=alt.cover or song.cover,
                    album=alt.album or song.album,
                    extra=alt.extra,
                )
                result = self._try_download_from_source(song, alt_song, save_dir)
                if result.success:
                    return result
                tried_sources.add(alt.source)
                if result.error:
                    errors.append(f"{alt.source}: {result.error}")
        except Exception:
            pass

        # ── 尝试 3: 全平台搜索，逐个尝试未试过的源 ──
        try:
            search_results = self.search_songs(
                f"{song.name} {song.artist}",
            )
            # 按匹配度排序：名字和歌手都匹配的优先
            candidates = []
            for s in search_results:
                if s.source in tried_sources:
                    continue
                # 简单匹配：歌名包含关系
                name_match = (
                    song.name.lower() in s.name.lower()
                    or s.name.lower() in song.name.lower()
                )
                if name_match:
                    candidates.append(s)

            for candidate in candidates:
                if candidate.source in tried_sources:
                    continue
                result = self._try_download_from_source(song, candidate, save_dir)
                tried_sources.add(candidate.source)
                if result.success:
                    return result
                if result.error:
                    errors.append(f"{candidate.source}: {result.error}")
        except Exception:
            pass

        # ── 全部穷尽 ──
        error_summary = f"已尝试 {len(tried_sources)} 个源均失败"
        if errors:
            # 只显示最后几个错误，避免太长
            last_errors = errors[-3:]
            error_summary += " (" + "; ".join(last_errors) + ")"
        return DownloadResult(
            song=song, success=False,
            actual_source=song.source,
            error=error_summary,
        )

    def _try_download_from_source(
        self,
        original_song: Song,
        download_song: Song,
        save_dir: str = None,
    ) -> "DownloadResult":
        """尝试从指定源下载单首歌曲。

        Args:
            original_song: 原始歌曲（用于文件名和歌词）
            download_song: 实际下载源的歌曲
            save_dir: 保存目录

        Returns:
            DownloadResult (success=True 表示完全成功)
        """
        # inspect 检查
        try:
            inspect_result = self.inspect(
                download_song.id, download_song.source,
                duration=download_song.duration,
            )
        except Exception:
            return DownloadResult(
                song=original_song, success=False,
                actual_source=download_song.source,
                error="inspect 失败",
            )

        if not inspect_result.valid:
            return DownloadResult(
                song=original_song, success=False,
                actual_source=download_song.source,
                error="资源不可用",
            )

        # 下载音频
        try:
            filepath = self.download(
                download_song.id, download_song.source,
                name=original_song.name, artist=original_song.artist,
                cover=download_song.cover, extra=download_song.extra,
                save_dir=save_dir,
            )
        except Exception as e:
            return DownloadResult(
                song=original_song, success=False,
                actual_source=download_song.source,
                error=f"下载失败: {e}",
            )

        # 下载歌词（尽力而为，不影响结果）
        lrc_path = None
        txt_path = None
        try:
            lrc_path = self.download_lyrics_file(
                original_song.id, original_song.source,
                name=original_song.name, artist=original_song.artist,
                save_dir=save_dir,
            )
            if lrc_path:
                txt_path = _save_txt_lyrics(lrc_path)
        except Exception:
            pass

        return DownloadResult(
            song=original_song, success=True,
            filepath=filepath, lrc_path=lrc_path, txt_path=txt_path,
            actual_source=download_song.source,
        )

    # ── Full Detail ────────────────────────────────────────────────────────

    def get_song_full_detail(self, song_id: str, source: str, duration: int = 0) -> Song:
        """获取歌曲完整详情 (inspect + lyrics)。

        组合调用 inspect 和 get_lyrics，返回带完整信息的 Song。

        Args:
            song_id: 歌曲 ID
            source: 音源平台
            duration: 歌曲时长

        Returns:
            带完整信息的 Song（含 url, size, bitrate, lyrics）
        """
        # Get resource info
        inspect_result = self.inspect(song_id, source, duration=duration)

        # Get lyrics
        lyrics = ""
        try:
            lyrics = self.get_lyrics(song_id, source)
        except Exception:
            pass

        return Song(
            id=song_id,
            source=source,
            name="",
            artist="",
            duration=duration,
            cover="",
            url=inspect_result.url,
            size=inspect_result.size,
            bitrate=inspect_result.bitrate,
            lyrics=lyrics,
        )

    def enrich_song(self, song: Song) -> Song:
        """为已有 Song 补充 inspect + lyrics 信息。

        Args:
            song: 基础 Song 对象（来自搜索结果）

        Returns:
            新 Song 对象，补充了 url/size/bitrate/lyrics
        """
        inspect_result = self.inspect(song.id, song.source, duration=song.duration)

        lyrics = ""
        try:
            lyrics = self.get_lyrics(song.id, song.source)
        except Exception:
            pass

        return Song(
            id=song.id,
            source=song.source,
            name=song.name,
            artist=song.artist,
            duration=song.duration,
            cover=song.cover,
            album=song.album,
            link=song.link,
            extra=song.extra,
            url=inspect_result.url,
            size=inspect_result.size,
            bitrate=inspect_result.bitrate,
            lyrics=lyrics,
        )

    # ── Platforms ──────────────────────────────────────────────────────────

    @staticmethod
    def list_platforms() -> dict[str, str]:
        """列出支持的平台。

        Returns:
            {平台代码: 中文名} 字典
        """
        return dict(PLATFORMS)

    # ── Cookie Management ──────────────────────────────────────────────────

    def get_cookies(self) -> dict[str, str]:
        """获取所有平台的 cookie 配置。

        Returns:
            {平台代码: cookie值} 字典
        """
        resp = self._get("/music/cookies")
        return resp.json()

    def set_cookies(self, cookies: dict[str, str]) -> None:
        """设置平台 cookie。

        Args:
            cookies: {平台代码: cookie值} 字典
                    例如: {"netease": "MUSIC_U=xxx", "qq": "uin=xxx; qm_keyst=xxx"}
                    传空字符串可删除对应平台的 cookie

        Example:
            client.set_cookies({
                "netease": "MUSIC_U=your_cookie_here",
                "qq": "uin=123456; qm_keyst=xxx",
                "kugou": ""  # 删除酷狗 cookie
            })
        """
        resp = self._session.post(
            f"{self.base_url}/music/cookies",
            json=cookies,
            timeout=self.timeout,
        )
        resp.raise_for_status()
        result = resp.json()
        if result.get("status") != "ok":
            raise MusicClientError(f"设置 cookie 失败: {result}")

    def set_cookie(self, source: str, cookie: str) -> None:
        """设置单个平台的 cookie。

        Args:
            source: 平台代码 (netease/qq/kugou/kuwo/migu/bilibili)
            cookie: cookie 字符串（传空字符串删除）

        Example:
            client.set_cookie("netease", "MUSIC_U=your_cookie_here")
        """
        self.set_cookies({source: cookie})

    def clear_cookies(self, sources: list[str] = None) -> None:
        """清除指定平台的 cookie。

        Args:
            sources: 平台列表，None 表示清除所有

        Example:
            client.clear_cookies(["netease", "qq"])  # 清除网易云和QQ音乐
            client.clear_cookies()  # 清除所有
        """
        if sources is None:
            # 清除所有：获取当前所有平台，设置为空
            current = self.get_cookies()
            cookies_to_clear = {k: "" for k in current.keys()}
        else:
            cookies_to_clear = {src: "" for src in sources}

        self.set_cookies(cookies_to_clear)

    def load_cookies_from_env(self) -> dict[str, str]:
        """从环境变量加载 cookie。

        环境变量命名规则:
            NETEASE_COOKIE, QQ_MUSIC_COOKIE, KUGOU_COOKIE,
            KUWO_COOKIE, MIGU_COOKIE, BILIBILI_COOKIE

        Returns:
            成功加载的 cookie 字典

        Example:
            export NETEASE_COOKIE="MUSIC_U=xxx"
            export QQ_MUSIC_COOKIE="uin=xxx; qm_keyst=xxx"

            client.load_cookies_from_env()
        """
        loaded = {}
        for source, env_var in COOKIE_ENV_VARS.items():
            cookie = os.environ.get(env_var, "").strip()
            if cookie:
                loaded[source] = cookie

        if loaded:
            self.set_cookies(loaded)

        return loaded

    def load_cookies_from_file(self, filepath: str = None) -> dict[str, str]:
        """从 JSON 文件加载 cookie。

        Args:
            filepath: JSON 文件路径，默认为 .music_cookies.json

        JSON 格式:
            {
                "netease": "MUSIC_U=xxx",
                "qq": "uin=xxx; qm_keyst=xxx",
                "kugou": "xxx"
            }

        Returns:
            成功加载的 cookie 字典

        Example:
            client.load_cookies_from_file(".music_cookies.json")
        """
        filepath = filepath or DEFAULT_COOKIE_FILE
        if not os.path.exists(filepath):
            return {}

        with open(filepath, "r", encoding="utf-8") as f:
            cookies = json.load(f)

        if cookies:
            self.set_cookies(cookies)

        return cookies

    def save_cookies_to_file(self, filepath: str = None) -> None:
        """保存当前 cookie 到 JSON 文件。

        Args:
            filepath: JSON 文件路径，默认为 .music_cookies.json

        Example:
            client.save_cookies_to_file(".music_cookies.json")
        """
        filepath = filepath or DEFAULT_COOKIE_FILE
        cookies = self.get_cookies()

        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(cookies, f, ensure_ascii=False, indent=2)

    # ── Internal HTTP ──────────────────────────────────────────────────────

    def _get(self, path: str, params: dict = None, stream: bool = False) -> requests.Response:
        """发起 GET 请求。"""
        url = f"{self.base_url}{path}"
        try:
            resp = self._session.get(url, params=params, timeout=self.timeout, stream=stream)
            resp.raise_for_status()
            return resp
        except requests.ConnectionError:
            raise MusicClientError(
                f"无法连接到 go-music-dl 服务: {self.base_url}\n"
                "请确保 Docker 容器正在运行。"
            )
        except requests.HTTPError as e:
            raise MusicClientError(f"HTTP 错误: {e}")
        except requests.Timeout:
            raise MusicClientError(f"请求超时 ({self.timeout}s): {url}")


# ─── Filename Helpers ─────────────────────────────────────────────────────────

def _extract_filename(resp: requests.Response) -> str:
    """从 Content-Disposition header 提取文件名。"""
    cd = resp.headers.get("Content-Disposition", "")
    if not cd:
        return ""
    from urllib.parse import unquote
    # Try filename*=UTF-8''... format first (case-insensitive)
    match = re.search(r"filename\*=(?:UTF-8|utf-8)''(.+?)(?:;|$)", cd)
    if match:
        return unquote(match.group(1).strip())
    # Then try filename="..." format
    match = re.search(r'filename="?([^";]+)"?', cd)
    if match:
        return unquote(match.group(1).strip())
    return ""


def _build_filename(name: str, artist: str, ext: str) -> str:
    """构建文件名: 歌名-歌手.ext"""
    if name and artist:
        return f"{name}-{artist}.{ext}"
    if name:
        return f"{name}.{ext}"
    return f"unknown.{ext}"


def _sanitize_filename(filename: str) -> str:
    """清理文件名中的不安全字符。"""
    # Replace unsafe chars with underscore
    return re.sub(r'[<>:"/\\|?*\x00-\x1f]', '_', filename).strip()


# ─── FeishuPusher: Feishu Integration ─────────────────────────────────────────

class FeishuPusher:
    """飞书推送器 — 将音乐数据推送到飞书群。

    依赖 feishu_toolkit.py（优先查找同级 `feishu-toolkit/` 目录）。

    Args:
        app_id: 飞书应用 ID
        app_secret: 飞书应用 Secret
        default_chat_id: 默认推送群组 ID
    """

    def __init__(
        self,
        app_id: str = None,
        app_secret: str = None,
        default_chat_id: str = None,
    ):
        self.default_chat_id = (
            default_chat_id
            or os.environ.get("FEISHU_DEFAULT_CHAT_ID", "")
        )

        # Import feishu_toolkit
        feishu_client_cls = _import_feishu_client()
        self._client = feishu_client_cls(
            app_id=app_id or os.environ.get("FEISHU_APP_ID", ""),
            app_secret=app_secret or os.environ.get("FEISHU_APP_SECRET", ""),
        )

    def _resolve_chat_id(self, chat_id: str = None) -> str:
        """解析 chat_id，优先使用传入值，否则用默认值。"""
        cid = chat_id or self.default_chat_id
        if not cid:
            raise ValueError(
                "chat_id 未指定。请传入 chat_id 参数或设置 FEISHU_DEFAULT_CHAT_ID 环境变量。"
            )
        return cid

    def push_song_card(self, song: Song, chat_id: str = None) -> dict:
        """推送单曲详情卡片到飞书群。

        卡片包含: 歌名、歌手、时长、音质、平台、歌词预览。

        Args:
            song: Song 对象
            chat_id: 目标群组 ID

        Returns:
            飞书 API 响应
        """
        cid = self._resolve_chat_id(chat_id)

        # Build fields
        fields = [
            (f"**歌手**: {song.artist}", True),
            (f"**时长**: {song.duration_str}", True),
        ]
        if song.album:
            fields.append((f"**专辑**: {song.album}", True))
        if song.bitrate:
            fields.append((f"**音质**: {song.bitrate}", True))
        if song.size:
            fields.append((f"**大小**: {song.size}", True))
        fields.append((f"**平台ID**: {song.id}", True))

        elements = [
            self._client.card_fields(fields),
        ]

        # Add lyrics preview if available
        if song.lyrics:
            lyrics_preview = _format_lyrics_preview(song.lyrics, max_lines=4)
            if lyrics_preview:
                elements.append(self._client.card_divider())
                elements.append(self._client.card_markdown(
                    f"📝 **歌词预览**\n{lyrics_preview}"
                ))

        # Source badge in title
        source_emoji = _source_emoji(song.source)
        title = f"🎵 {song.name}    {source_emoji} {song.source_name}"

        card = self._client.build_card(title, elements, color="blue")
        return self._client.send_card(cid, card)

    def push_search_results(
        self,
        songs: list[Song],
        keyword: str,
        chat_id: str = None,
    ) -> dict:
        """推送搜索结果列表卡片到飞书群。

        Args:
            songs: 歌曲列表
            keyword: 搜索关键词
            chat_id: 目标群组 ID

        Returns:
            飞书 API 响应
        """
        cid = self._resolve_chat_id(chat_id)

        lines = []
        for i, song in enumerate(songs[:10], 1):
            line = (
                f"{i}. **{song.name}** - {song.artist} "
                f"({song.duration_str}) [{song.source_name}]"
            )
            if song.size:
                line += f" {song.size}"
            lines.append(line)

        body_text = "\n".join(lines)
        if len(songs) > 10:
            body_text += f"\n\n... 共 {len(songs)} 首"

        elements = [
            self._client.card_markdown(body_text),
        ]

        title = f"🔍 搜索: {keyword} ({len(songs)} 首)"
        card = self._client.build_card(title, elements, color="green")
        return self._client.send_card(cid, card)

    def push_playlist_card(
        self,
        playlist: Playlist,
        songs: list[Song] = None,
        chat_id: str = None,
    ) -> dict:
        """推送歌单卡片到飞书群。

        Args:
            playlist: 歌单对象
            songs: 歌单中的歌曲列表 (可选，显示前 10 首)
            chat_id: 目标群组 ID

        Returns:
            飞书 API 响应
        """
        cid = self._resolve_chat_id(chat_id)

        fields = [
            (f"**创建者**: {playlist.creator}", True),
            (f"**歌曲数**: {playlist.track_count}", True),
            (f"**平台**: {playlist.source_name}", True),
        ]
        if playlist.play_count > 0:
            fields.append((f"**播放量**: {playlist.play_count:,}", True))

        elements = [self._client.card_fields(fields)]

        if playlist.description:
            elements.append(self._client.card_markdown(playlist.description))

        if songs:
            elements.append(self._client.card_divider())
            song_lines = []
            for i, song in enumerate(songs[:10], 1):
                song_lines.append(
                    f"{i}. **{song.name}** - {song.artist} ({song.duration_str})"
                )
            if len(songs) > 10:
                song_lines.append(f"... 共 {len(songs)} 首")
            elements.append(self._client.card_markdown("\n".join(song_lines)))

        title = f"📋 {playlist.name}"
        card = self._client.build_card(title, elements, color="purple")
        return self._client.send_card(cid, card)

    def push_playlist_detail_card(
        self,
        playlist: "PlaylistDetailInfo",
        chat_id: str = None,
        max_tracks: int = 0,
        sort_by: str = "",
        sort_desc: bool = True,
        with_doc: bool = False,
    ) -> dict:
        """推送歌单详情卡片（含曲目统计数据）到飞书群。

        卡片结构:
          - Header: 歌单标题  副标题: 创建者 · N 首
          - Fields: 更新日期、收藏、分享数
          - 曲目列表: 单 markdown 文本块（轻量，支持 100+ 首）
            格式: 序号. 歌名(链接) — 歌手  👍收藏  💬评论  🔗分享
          - 底部备注

        with_doc=True 时，推送卡片后额外生成 CSV 文件发送到同群，
        包含所有曲目的完整数据。

        Args:
            playlist: PlaylistDetailInfo 对象
            chat_id: 目标群组 ID
            max_tracks: 最多显示曲目数（0 = 全部）
            sort_by: 排序字段 likes / comments / shares / date（空 = 歌单原序）
            sort_desc: True=降序 False=升序
            with_doc: 推送后生成 CSV 并发到群

        Returns:
            飞书 API 响应
        """
        import datetime as _dt

        c = self._client
        cid = self._resolve_chat_id(chat_id)

        # ── 排序 ──
        _sort_keys = {
            "likes":    lambda t: t.favorites,
            "comments": lambda t: t.comments,
            "shares":   lambda t: t.shares,
            "date":     lambda t: t.publish_date or "0000-00-00",
        }
        tracks = list(playlist.tracks)
        if sort_by in _sort_keys:
            tracks.sort(key=_sort_keys[sort_by], reverse=sort_desc)

        display_tracks = tracks if not max_tracks else tracks[:max_tracks]

        # ── 数字格式化（万为单位） ──
        def _fmt(n: int) -> str:
            if n >= 10000:
                return f"{n / 10000:.1f}万"
            return f"{n:,}"

        # ── 歌单基本信息（fields 双列） ──
        fields = []
        if playlist.update_time:
            fields.append((f"**更新时间**\n{playlist.update_time}", True))
        if playlist.create_time:
            fields.append((f"**创建时间**\n{playlist.create_time}", True))
        if playlist.extra.get("collect_count"):
            fields.append((f"**收藏**\n{playlist.extra['collect_count']:,}", True))
        if playlist.extra.get("share_count"):
            fields.append((f"**分享**\n{playlist.extra['share_count']:,}", True))
        if playlist.extra.get("play_count"):
            fields.append((f"**播放**\n{playlist.extra['play_count']:,}", True))

        elements = []
        if fields:
            elements.append(c.card_fields(fields))

        # ── 曲目列表：单个 2 列 column_set（名称 | 数据，全部行一次对齐） ──
        if display_tracks:
            elements.append(c.card_divider())

            show_date = sort_by == "date"

            # 列标题行（纯文字 + 颜色标签，无表情）
            stat_parts = []
            if show_date:
                stat_parts.append(c.md_tag("日期", "orange") if sort_by == "date" else "日期")
            stat_parts.append(c.md_tag("收藏", "red") if sort_by == "likes" else "收藏")
            stat_parts.append(c.md_tag("评论", "turquoise") if sort_by == "comments" else "评论")
            stat_parts.append(c.md_tag("分享", "violet") if sort_by == "shares" else "分享")
            stat_header = "　　".join(stat_parts)

            elements.append(c.card_column_set(
                c.card_column([c.card_markdown("**歌曲名 — 歌手**")], weight=6),
                c.card_column([c.card_markdown(stat_header)], weight=4),
            ))

            # 全部行合并到一个 column_set 的两个 markdown 块
            name_lines = []
            stat_lines = []
            for i, t in enumerate(display_tracks, 1):
                link = t.resolved_url or (
                    f"https://music.douyin.com/qishui/share/track?track_id={t.song_id}"
                    if t.song_id else ""
                )
                name = f"[{t.name}]({link})" if link else t.name
                name_lines.append(f"{i}. **{name}** — {t.artist}")

                if show_date:
                    stat_lines.append(
                        f"{t.publish_date or '—'}　　"
                        f"{_fmt(t.favorites)}　　{_fmt(t.comments)}　　{_fmt(t.shares)}"
                    )
                else:
                    stat_lines.append(
                        f"{_fmt(t.favorites)}　　{_fmt(t.comments)}　　{_fmt(t.shares)}"
                    )

            elements.append(c.card_column_set(
                c.card_column([c.card_markdown("\n".join(name_lines))], weight=6),
                c.card_column([c.card_markdown("\n".join(stat_lines))], weight=4),
            ))

        # ── 底部备注 ──
        sort_label_text = {
            "likes": "点赞", "comments": "评论",
            "shares": "分享", "date": "日期",
        }
        sort_info = ""
        if sort_by in sort_label_text:
            direction = "降序" if sort_desc else "升序"
            sort_info = f"  ·  排序: {sort_label_text[sort_by]} {direction}"

        cap_info = ""
        if max_tracks and len(tracks) > max_tracks:
            cap_info = f"  ·  显示前 {max_tracks} 首，共 {len(tracks)} 首"

        today = _dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        doc_hint = "  |  完整数据见下方文件" if with_doc else ""
        elements.append(c.card_note(
            c.note_md(
                f"数据来源: {playlist.source_name}  |  抓取于 {today}"
                f"{sort_info}{cap_info}{doc_hint}"
            )
        ))

        subtitle = f"{playlist.creator}  ·  {len(playlist.tracks)} 首"
        card = c.build_card(playlist.title, elements, color="wathet", subtitle=subtitle)
        result = c.send_card(cid, card)

        # ── 生成文件并发送到群 ──
        if with_doc:
            now_str = _dt.datetime.now().strftime("%Y%m%d_%H%M%S")
            self._send_playlist_csv(playlist, tracks, cid, now_str)

        return result

    def create_playlist_bitable(
        self,
        playlist: "PlaylistDetailInfo",
        tracks: list = None,
        sort_by: str = "",
        sort_desc: bool = True,
    ) -> str:
        """创建歌单多维表格，返回 bitable URL。

        字段: 序号 / 平台 / song_id / 歌名 / 歌手 / 时长 / 专辑 / 发布日期 /
              收藏 / 评论 / 分享 / 播放 / 链接

        支持汽水、网易云、QQ音乐等所有平台，字段统一，无数据的列值为 0。

        Args:
            playlist:  PlaylistDetailInfo 对象
            tracks:    已排序曲目列表（None 则按 sort_by 自动排序）
            sort_by:   likes / comments / shares / date（空 = 歌单原序）
            sort_desc: True=降序  False=升序

        Returns:
            飞书多维表格 URL
        """
        import datetime as _dt
        import re as _re

        c = self._client

        if tracks is None:
            _sort_keys = {
                "likes":    lambda t: t.favorites,
                "comments": lambda t: t.comments,
                "shares":   lambda t: t.shares,
                "date":     lambda t: t.publish_date or "0000-00-00",
            }
            tracks = list(playlist.tracks)
            if sort_by in _sort_keys:
                tracks.sort(key=_sort_keys[sort_by], reverse=sort_desc)

        date_str = _dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_title = _re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", playlist.title)[:40]

        field_defs = [
            ("序号",    1),
            ("平台",    1),
            ("song_id", 1),
            ("歌名",    1),
            ("歌手",    1),
            ("时长",    1),
            ("专辑",    1),
            ("发布日期", 1),
            ("收藏",    2),
            ("评论",    2),
            ("分享",    2),
            ("播放",    2),
            ("链接",    1),
        ]
        app_token, table_id, bitable_url = c.create_bitable_with_fields(
            f"{safe_title}_{date_str}", field_defs,
        )

        platform_name = playlist.source_name
        records = []
        for i, t in enumerate(tracks, 1):
            records.append({"fields": {
                "序号":    str(i),
                "平台":    platform_name,
                "song_id": t.song_id,
                "歌名":    t.name,
                "歌手":    t.artist,
                "时长":    t.duration_str,
                "专辑":    t.album,
                "发布日期": t.publish_date,
                "收藏":    t.favorites,
                "评论":    t.comments,
                "分享":    t.shares,
                "播放":    t.plays,
                "链接":    t.resolved_url,
            }})

        for j in range(0, len(records), 500):
            c.create_bitable_records(app_token, table_id, records[j:j + 500])

        try:
            c.set_drive_public_permission(app_token, file_type="bitable")
        except Exception:
            pass

        return bitable_url

    def _send_playlist_csv(
        self,
        playlist: "PlaylistDetailInfo",
        tracks: list,
        chat_id: str,
        date_str: str,
    ) -> None:
        """创建飞书在线表格 + 发送 CSV/XLSX 文件到群。

        三种输出，一次完成：
          1. 飞书多维表格（在线查看、排序、筛选、团队共享）
          2. CSV 文件（通用，任何软件可打开）
          3. XLSX 文件（Excel 原生格式，有列宽和表头样式）
        """
        import csv
        import tempfile
        from pathlib import Path

        c = self._client
        safe_title = re.sub(r'[<>:"/\\|?*\x00-\x1f]', '_', playlist.title)[:40]
        bitable_name = f"{safe_title}_{date_str}"

        # ── 构建行数据（CSV/XLSX 共享，数字全部转文本避免科学计数法） ──
        headers = ["序号", "平台", "song_id", "歌名", "歌手", "时长", "专辑",
                   "发布日期", "收藏", "评论", "分享", "播放", "链接"]
        platform_name = playlist.source_name
        rows = []
        for i, t in enumerate(tracks, 1):
            rows.append([
                str(i), platform_name, str(t.song_id), t.name, t.artist,
                t.duration_str, t.album, t.publish_date,
                str(t.favorites), str(t.comments), str(t.shares),
                str(t.plays), t.resolved_url,
            ])

        tmp_dir = Path(tempfile.gettempdir())

        # ── 1. 飞书多维表格 ──
        bitable_url = ""
        try:
            bitable_url = self.create_playlist_bitable(playlist, tracks=tracks)
            print(f"   [在线表格] {bitable_url}")
        except Exception as e:
            print(f"   ⚠️  在线表格创建失败: {e}", file=__import__("sys").stderr)

        # ── 2. CSV 文件 ──
        csv_path = tmp_dir / f"{safe_title}_{date_str}.csv"
        try:
            with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(headers)
                writer.writerows(rows)
            file_key = c.upload_file(str(csv_path))
            c.send_file(chat_id, file_key)
            print(f"   [CSV] {csv_path.name}")
        except Exception as e:
            print(f"   ⚠️  CSV 失败: {e}", file=__import__("sys").stderr)
        finally:
            csv_path.unlink(missing_ok=True)

        # ── 3. XLSX 文件（需要 openpyxl，没有则跳过） ──
        xlsx_path = tmp_dir / f"{safe_title}_{date_str}.xlsx"
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill

            wb = Workbook()
            ws = wb.active
            ws.title = playlist.title[:30] or "歌单数据"

            # 表头样式
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            # 数据行（XLSX 用原始数字类型，不会科学计数法）
            # 新列顺序: 序号(0) 平台(1) song_id(2) 歌名(3) 歌手(4) 时长(5) 专辑(6)
            #           发布日期(7) 收藏(8) 评论(9) 分享(10) 播放(11) 链接(12)
            _num_cols = {8, 9, 10, 11}  # 收藏、评论、分享、播放的列索引(0-based)
            for row_idx, row_data in enumerate(rows, 2):
                for col_idx, val in enumerate(row_data, 1):
                    # 数字列转回 int 给 Excel 用
                    if (col_idx - 1) in _num_cols and isinstance(val, str) and val.isdigit():
                        val = int(val)
                    ws.cell(row=row_idx, column=col_idx, value=val)

            # 自动列宽（按内容估算）
            col_widths = {
                "序号": 6, "平台": 12, "song_id": 22, "歌名": 25, "歌手": 15,
                "时长": 8, "专辑": 20, "发布日期": 12,
                "收藏": 10, "评论": 10, "分享": 10, "播放": 10, "链接": 40,
            }
            for col, h in enumerate(headers, 1):
                ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = col_widths.get(h, 12)

            # 冻结首行
            ws.freeze_panes = "A2"

            wb.save(str(xlsx_path))
            file_key = c.upload_file(str(xlsx_path))
            c.send_file(chat_id, file_key)
            print(f"   [XLSX] {xlsx_path.name}")
        except ImportError:
            pass  # openpyxl 未安装，静默跳过
        except Exception as e:
            print(f"   ⚠️  XLSX 失败: {e}", file=__import__("sys").stderr)
        finally:
            xlsx_path.unlink(missing_ok=True)

        # ── 发送汇总卡片（含在线表格链接） ──
        elements = [
            c.card_markdown(
                f"共 **{len(tracks)}** 首曲目数据已导出\n"
                f"CSV + XLSX 文件见上方\n"
                f"在线表格支持排序、筛选、团队共享"
            ),
        ]
        if bitable_url:
            elements.append(c.card_button("打开在线表格", bitable_url))
        card = c.build_card(
            f"{playlist.title} — 完整数据",
            elements,
            color="green",
        )
        c.send_card(chat_id, card)

    def create_song_document(self, song: Song) -> str:
        """创建飞书文档记录歌曲详情。

        Args:
            song: Song 对象（建议先 enrich）

        Returns:
            文档 URL
        """
        blocks = [
            self._client.heading_block(f"🎵 {song.name}", level=1),
            self._client.text_block(
                f"歌手: {song.artist}  |  时长: {song.duration_str}  |  平台: {song.source_name}"
            ),
            self._client.divider_block(),
        ]

        if song.album:
            blocks.append(self._client.text_block(f"专辑: {song.album}"))
        if song.bitrate:
            blocks.append(self._client.text_block(f"音质: {song.bitrate}  |  大小: {song.size}"))

        if song.lyrics:
            blocks.append(self._client.heading_block("歌词", level=2))
            # Split lyrics into reasonable blocks
            for chunk in _split_lyrics_for_doc(song.lyrics):
                blocks.append(self._client.text_block(chunk))

        result = self._client.create_document_with_content(
            title=f"{song.name} - {song.artist}",
            blocks=blocks,
        )
        return result.get("url", "")

    def send_song_files(
        self,
        file_paths: list[Path],
        chat_id: str = None,
        zip_name: str = "",
    ) -> dict:
        """发送歌曲文件到飞书群。

        逻辑:
          - 1 个文件且 ≤30MB: 直接上传到 IM 并发送
          - 多个文件: 打包成 1 个 zip
            - zip ≤30MB: 通过 IM 上传发送
            - zip >30MB: 通过 Drive 分片上传 → 发送卡片含下载链接

        Args:
            file_paths: 文件路径列表 (mp3/m4a/flac/lrc 等)
            chat_id: 目标群组 ID
            zip_name: 压缩包文件名 (多文件时使用，默认自动生成)

        Returns:
            飞书 API 响应
        """
        import zipfile
        import tempfile

        MAX_IM_FILE_SIZE = 30 * 1024 * 1024  # 30MB IM 上传限制

        cid = self._resolve_chat_id(chat_id)

        if not file_paths:
            raise ValueError("file_paths 不能为空")

        # 过滤掉不存在的文件
        existing = [Path(p) for p in file_paths if Path(p).exists()]
        if not existing:
            raise FileNotFoundError("所有文件均不存在")

        # ── 快速路径: 单个小文件直接发送 ──
        if len(existing) == 1 and existing[0].stat().st_size <= MAX_IM_FILE_SIZE:
            file_key = self._client.upload_file(str(existing[0]))
            return self._client.send_file(cid, file_key)

        # ── 多文件或大文件: 打包成 1 个 zip ──
        if not zip_name:
            zip_name = "songs"
        base_name = zip_name.removesuffix(".zip")
        zip_filename = f"{base_name}.zip"

        with tempfile.TemporaryDirectory() as tmp_dir:
            zip_path = Path(tmp_dir) / _sanitize_filename(zip_filename)
            with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
                for fp in existing:
                    zf.write(fp, fp.name)

            zip_size = zip_path.stat().st_size
            zip_size_mb = zip_size / (1024 * 1024)

            if zip_size <= MAX_IM_FILE_SIZE:
                # zip ≤30MB: 通过 IM 直接发送
                print(f"   📦 上传: {zip_filename} ({zip_size_mb:.1f} MB)")
                file_key = self._client.upload_file(str(zip_path))
                return self._client.send_file(cid, file_key)

            # zip >30MB: 通过 Drive 分片上传
            print(f"   📦 打包完成: {zip_filename} ({zip_size_mb:.1f} MB)")
            print(f"   ☁️  使用云盘分片上传...")

            return self._upload_to_drive_and_share(
                zip_path, zip_filename, zip_size_mb, cid,
            )

    def _upload_to_drive_and_share(
        self,
        file_path: Path,
        display_name: str,
        size_mb: float,
        chat_id: str,
    ) -> dict:
        """通过 Drive 分片上传文件并发送卡片消息到群。

        Args:
            file_path: 本地文件路径
            display_name: 显示名称
            size_mb: 文件大小 (MB)
            chat_id: 飞书群 ID

        Returns:
            飞书 API 响应
        """
        # 获取/创建 Music 文件夹
        root_token = self._client.get_root_folder_token()
        music_folder = self._client.find_or_create_folder("Music", root_token)

        # 分片上传
        def _on_progress(seq, total):
            print(f"   ⏫ 上传分片 [{seq}/{total}]")

        file_token = self._client.upload_file_to_drive(
            str(file_path),
            parent_node=music_folder,
            on_progress=_on_progress,
        )
        print(f"   ✅ 上传完成: file_token={file_token}")

        # 设置分享权限 (组织内可读)
        try:
            self._client.set_drive_public_permission(file_token, file_type="file")
        except Exception as e:
            print(f"   ⚠️  设置分享权限失败 (可手动分享): {e}")

        # 构建下载链接
        file_url = f"https://feishu.cn/file/{file_token}"

        # 发送卡片消息到群
        card = self._client.build_card(
            f"📦 {display_name}",
            [
                self._client.card_fields([
                    (f"**文件名**: {display_name}", True),
                    (f"**大小**: {size_mb:.1f} MB", True),
                ]),
                self._client.card_button("📥 下载文件", file_url),
            ],
            color="blue",
        )
        return self._client.send_card(chat_id, card)

    def create_playlist_lyrics_doc(
        self,
        songs: list[Song],
        title: str = "歌单歌词",
    ) -> str:
        """创建飞书文档记录歌单所有歌词。

        每首歌名作为 H1 标题，歌词内容在标题下方，方便查找。

        Args:
            songs: 歌曲列表（需含 lyrics 字段）
            title: 文档标题

        Returns:
            文档 URL
        """
        blocks = []
        for song in songs:
            # 歌名 H1
            song_title = f"{song.name} - {song.artist}"
            blocks.append(self._client.heading_block(song_title, level=1))

            # 歌曲信息
            meta = f"时长: {song.duration_str}  |  平台: {song.source_name}"
            if song.album:
                meta += f"  |  专辑: {song.album}"
            blocks.append(self._client.text_block(meta))

            # 歌词内容（代码块，方便一键复制）
            if song.lyrics:
                lyrics_text = _lrc_to_text(song.lyrics)
                if lyrics_text:
                    blocks.append(self._client.code_block(lyrics_text))
                else:
                    blocks.append(self._client.text_block("(暂无歌词)"))
            else:
                blocks.append(self._client.text_block("(暂无歌词)"))

            # 分割线
            blocks.append(self._client.divider_block())

        result = self._client.create_document_with_content(
            title=title,
            blocks=blocks,
        )

        doc_url = result.get("url", "")

        # 设置文档分享权限
        doc_id = result.get("document_id", "")
        if doc_id:
            try:
                self._client.set_drive_public_permission(
                    doc_id, file_type="docx",
                )
            except Exception:
                pass

        return doc_url


# ─── Webhook Pusher (无需认证) ───────────────────────────────────────────────

def _lrc_to_text(lyrics: str) -> str:
    """将 LRC 歌词转换为纯文本（去掉时间标签和元数据行）。"""
    lines = []
    for line in lyrics.splitlines():
        text = re.sub(r'\[\d{2}:\d{2}[\.\d]*\]', '', line).strip()
        if text and not re.match(r'^\[.+\]$', text):
            lines.append(text)
    return "\n".join(lines)


def _save_txt_lyrics(lrc_path: Path) -> Optional[Path]:
    """读取 .lrc 歌词文件，转换为纯文本（去掉时间轴），保存为 .txt。"""
    try:
        lrc_content = lrc_path.read_text(encoding="utf-8")
        plain = _lrc_to_text(lrc_content)
        if not plain.strip():
            return None
        txt_path = lrc_path.with_suffix(".txt")
        txt_path.write_text(plain, encoding="utf-8")
        return txt_path
    except Exception:
        return None


def _send_webhook(webhook_url: str, payload: dict) -> dict:
    """发送单条 webhook 消息。"""
    resp = requests.post(webhook_url, json=payload, timeout=10)
    resp.raise_for_status()
    return resp.json()


def push_to_webhook(webhook_url: str, song: Song) -> dict:
    """推送歌曲详情卡片 + 纯文本歌词到飞书 webhook。

    发送两条消息:
      1. 交互卡片: 歌名、歌手、时长、平台、音质等
      2. 纯文本: 完整歌词（LRC 转纯文本）

    Args:
        webhook_url: 飞书 webhook URL
        song: Song 对象（建议先 enrich 获取完整信息）

    Returns:
        卡片消息的响应 JSON
    """
    # ── 1. 歌曲详情卡片 ──
    fields = [
        {"is_short": True, "text": {"tag": "lark_md", "content": f"**歌手**: {song.artist}"}},
        {"is_short": True, "text": {"tag": "lark_md", "content": f"**时长**: {song.duration_str}"}},
        {"is_short": True, "text": {"tag": "lark_md", "content": f"**平台**: {song.source_name}"}},
        {"is_short": True, "text": {"tag": "lark_md", "content": f"**ID**: {song.id}"}},
    ]
    if song.album:
        fields.append({"is_short": True, "text": {"tag": "lark_md", "content": f"**专辑**: {song.album}"}})
    if song.size:
        fields.append({"is_short": True, "text": {"tag": "lark_md", "content": f"**大小**: {song.size}"}})
    if song.bitrate and song.bitrate != "-":
        fields.append({"is_short": True, "text": {"tag": "lark_md", "content": f"**音质**: {song.bitrate}"}})

    elements = [{"tag": "div", "fields": fields}]

    # 卡片内加歌词预览（前 4 行）
    if song.lyrics:
        preview = _format_lyrics_preview(song.lyrics, max_lines=4)
        if preview:
            elements.append({"tag": "hr"})
            elements.append({
                "tag": "div",
                "text": {"tag": "lark_md", "content": f"📝 **歌词预览**\n{preview}"},
            })

    source_emoji = _source_emoji(song.source)
    card_payload = {
        "msg_type": "interactive",
        "card": {
            "header": {
                "title": {"tag": "plain_text", "content": f"🎵 {song.name}  {source_emoji} {song.source_name}"},
                "template": "blue",
            },
            "elements": elements,
        },
    }

    card_result = _send_webhook(webhook_url, card_payload)

    # ── 2. 完整歌词（纯文本消息） ──
    if song.lyrics:
        lyrics_text = _lrc_to_text(song.lyrics)
        if lyrics_text:
            text_payload = {
                "msg_type": "text",
                "content": {"text": f"📝 {song.name} - {song.artist} 完整歌词\n\n{lyrics_text}"},
            }
            _send_webhook(webhook_url, text_payload)

    return card_result


def _push_playlist_download_report(
    webhook_url: str,
    playlist_id: str,
    source: str,
    results: list,
    songs: list,
) -> dict:
    """推送歌单批量下载报告到飞书 webhook。

    Args:
        webhook_url: 飞书 webhook URL
        playlist_id: 歌单 ID
        source: 原始平台
        results: DownloadResult 列表
        songs: 原始歌曲列表

    Returns:
        飞书 API 响应
    """
    success = [r for r in results if r.success]
    failed = [r for r in results if not r.success]
    switched = [r for r in success if r.actual_source != source]

    # 构建歌曲列表文本
    lines = []
    for i, r in enumerate(results, 1):
        if r.success:
            src_tag = ""
            if r.actual_source != source:
                src_name = PLATFORMS.get(r.actual_source, r.actual_source)
                src_tag = f" 🔄{src_name}"
            lines.append(f"{i}. ✅ **{r.song.name}** - {r.song.artist}{src_tag}")
        else:
            lines.append(f"{i}. ❌ **{r.song.name}** - {r.song.artist} ({r.error})")

    body_text = "\n".join(lines)

    # 统计字段
    fields = [
        {"is_short": True, "text": {"tag": "lark_md", "content": f"**歌曲数**: {len(results)}"}},
        {"is_short": True, "text": {"tag": "lark_md", "content": f"**成功**: {len(success)}"}},
        {"is_short": True, "text": {"tag": "lark_md", "content": f"**失败**: {len(failed)}"}},
        {"is_short": True, "text": {"tag": "lark_md", "content": f"**换源**: {len(switched)}"}},
    ]

    source_name = PLATFORMS.get(source, source)
    elements = [
        {"tag": "div", "fields": fields},
        {"tag": "hr"},
        {"tag": "div", "text": {"tag": "lark_md", "content": body_text}},
    ]

    card_payload = {
        "msg_type": "interactive",
        "card": {
            "header": {
                "title": {"tag": "plain_text", "content": f"📋 歌单下载完成 ({len(success)}/{len(results)})"},
                "template": "green" if not failed else "orange",
            },
            "elements": elements,
        },
    }

    result = _send_webhook(webhook_url, card_payload)
    print(f"✅ 下载报告已推送到飞书")
    return result


# ─── Feishu Import Helper ────────────────────────────────────────────────────

def _import_feishu_client():
    """动态导入 FeishuClient"""
    env_path = os.environ.get("FEISHU_TOOLKIT_PATH", "").strip()
    project_root = os.path.abspath(os.path.dirname(__file__))
    workspace_root = os.path.abspath(os.path.join(project_root, ".."))
    feishu_paths = [
        env_path,
        os.path.join(workspace_root, "feishu-toolkit"),
        os.path.join(os.getcwd(), "feishu-toolkit"),
        os.path.join(os.path.abspath(os.path.join(os.getcwd(), "..")), "feishu-toolkit"),
        os.path.expanduser("~/claude-workspaces/feishu-toolkit"),
        os.path.expanduser("~/feishu-toolkit"),
    ]
    for p in feishu_paths:
        if not p:
            continue
        p = os.path.abspath(os.path.expanduser(p))
        if os.path.isfile(os.path.join(p, "feishu_toolkit.py")):
            if p not in sys.path:
                sys.path.insert(0, p)
            break

    try:
        from feishu_toolkit import FeishuClient
        return FeishuClient
    except ImportError:
        raise ImportError(
            "feishu_toolkit.py 未找到。请确保 feishu-toolkit 与 music-toolkit 位于同一工作区目录，\n"
            "或设置 FEISHU_TOOLKIT_PATH 指向 feishu-toolkit 所在目录。\n"
            "兼容旧布局时，也可继续使用 ~/feishu-toolkit。\n"
            "飞书推送功能需要 feishu-toolkit 支持。"
        )


# ─── Music Detail Scraper ────────────────────────────────────────────────────

_SCRAPER_MOBILE_UA = (
    "Mozilla/5.0 (iPhone; CPU iPhone OS 16_6 like Mac OS X) "
    "AppleWebKit/605.1.15 (KHTML, like Gecko) "
    "Version/16.6 Mobile/15E148 Safari/604.1"
)

_SCRAPER_PC_UA = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/120.0.0.0 Safari/537.36"
)

# endMs sentinel value Qishui uses for the last (credit) lyric line
_SODA_LYRICS_SENTINEL_MS = 9_007_199_254_740_991


def _soda_lyrics_to_lrc(lyrics_data: dict) -> str:
    """将汽水音乐歌词 JSON 转换为 LRC 格式。

    过滤掉 "作曲/作词" 元数据行和末尾的贡献者行（endMs 为 sentinel）。
    """
    if not lyrics_data:
        return ""
    sentences = lyrics_data.get("sentences", [])
    lines = []
    for s in sentences:
        start_ms = int(s.get("startMs") or 0)
        end_ms = int(s.get("endMs") or 0)
        text = (s.get("text") or "").strip()
        if not text:
            continue
        # Skip sentinel credit line at the very end
        if end_ms >= _SODA_LYRICS_SENTINEL_MS:
            continue
        minutes = start_ms // 60000
        seconds = (start_ms % 60000) / 1000
        lines.append(f"[{minutes:02d}:{seconds:05.2f}]{text}")
    return "\n".join(lines)


def _soda_format_qualities(bit_rates: list) -> str:
    """将 bit_rates 列表格式化为人类可读的音质字符串。

    示例: "medium(68k) / higher(132k) / highest(260k) / lossless(684k)"
    """
    if not bit_rates:
        return ""
    # Sort by bitrate ascending
    ordered = sorted(bit_rates, key=lambda b: int(b.get("br") or 0))
    parts = []
    for b in ordered:
        quality = b.get("quality") or ""
        br = int(b.get("br") or 0)
        if br:
            parts.append(f"{quality}({br // 1000}k)")
        else:
            parts.append(quality)
    return " / ".join(parts)


def _detect_platform_from_url(url: str) -> str:
    """从 URL 检测音乐平台代码。"""
    u = url.lower()
    if "qishui.douyin.com" in u or "music.douyin.com" in u:
        return "soda"
    if "music.163.com" in u or "163cn.tv" in u:
        return "netease"
    if "y.qq.com" in u or "c6.y.qq.com" in u:
        return "qq"
    if "kugou.com" in u:
        return "kugou"
    if "kuwo.cn" in u:
        return "kuwo"
    if "bilibili.com" in u or "b23.tv" in u:
        return "bilibili"
    return "unknown"


def _scrape_soda_detail(share_url: str, timeout: int = 15) -> SongDetailInfo:
    """抓取汽水音乐（SODA/Qishui）歌曲详情。

    支持:
      - 短链: https://qishui.douyin.com/s/CODE/
      - 完整: https://qishui.douyin.com/music/detail/MUSIC_ID
      - 分享: https://music.douyin.com/qishui/share/track?track_id=...

    数据来源: 页面内嵌 _ROUTER_DATA → loaderData → track_page → audioWithLyricsOption
    """
    import datetime

    session = requests.Session()
    session.headers.update({
        "User-Agent": _SCRAPER_MOBILE_UA,
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
    })

    # 1. Follow redirect → get page HTML and final URL
    resp = session.get(share_url, timeout=timeout, allow_redirects=True)
    resp.raise_for_status()
    final_url = resp.url
    page_html = resp.text

    # 2. Extract _ROUTER_DATA = {...}; function ...
    m = re.search(
        r'_ROUTER_DATA\s*=\s*(\{.+?\});\s*(?:function|window)',
        page_html,
        re.DOTALL,
    )
    if not m:
        raise ValueError(
            f"无法从页面提取歌曲数据 (未找到 _ROUTER_DATA)\nURL: {final_url}"
        )

    router_data = json.loads(m.group(1))
    track_page = router_data.get("loaderData", {}).get("track_page", {})
    audio = track_page.get("audioWithLyricsOption", {})
    track_info = audio.get("trackInfo", {})

    if not audio:
        raise ValueError(
            f"_ROUTER_DATA 中未找到 audioWithLyricsOption\nURL: {final_url}"
        )

    # 3. Basic info
    song_id = str(audio.get("track_id") or track_info.get("id") or "")
    name = audio.get("trackName") or track_info.get("name") or ""

    # Artist
    artist = audio.get("artistName") or ""
    if not artist:
        artists_list = track_info.get("artists") or []
        artist = "、".join(a.get("name", "") for a in artists_list if a.get("name"))

    # Duration: audio.duration is float seconds; trackInfo.duration is ms
    duration_float = float(audio.get("duration") or 0)
    if not duration_float:
        duration_float = float(track_info.get("duration") or 0) / 1000
    duration = int(duration_float)

    # 4. Cover — audio.coverURL is already a ready-to-use URL
    cover = audio.get("coverURL") or ""
    if not cover:
        album_info_tmp = track_info.get("album", {})
        url_cover = album_info_tmp.get("url_cover", {})
        url_bases = url_cover.get("urls", [])
        uri = url_cover.get("uri", "")
        tpl = url_cover.get("template_prefix", "")
        if url_bases and uri:
            cover = f"{url_bases[0]}{uri}~{tpl}-image.webp"

    # 5. Album
    album_info = track_info.get("album", {})
    album = album_info.get("name") or ""
    album_id = str(audio.get("album_id") or album_info.get("id") or "")

    # 6. Engagement stats
    stats = track_info.get("stats", {})
    favorites = int(stats.get("count_collected") or 0)
    comments = int(
        stats.get("count_comment")
        or audio.get("commentsStruct", {}).get("count")
        or 0
    )
    shares = int(stats.get("count_shared") or 0)
    plays = int(stats.get("count_played") or stats.get("count_view") or 0)

    # 7. Publish date (prefer album release_date, fall back to track create_time)
    create_time = int(album_info.get("release_date") or audio.get("create_time") or 0)
    publish_date = ""
    if create_time:
        publish_date = datetime.datetime.fromtimestamp(create_time).strftime("%Y-%m-%d")

    # 8. Audio CDN URL (temporary, ~24h TTL)
    audio_url = audio.get("url") or ""

    # 9. Lyrics → LRC format
    lyrics_lrc = _soda_lyrics_to_lrc(audio.get("lyrics") or {})

    # 10. Genre & language
    genre = audio.get("genre_tag") or ""
    lang_codes = track_info.get("lang_codes") or []
    language = ", ".join(lang_codes)

    # 11. Composers & lyricists
    smt = track_info.get("song_maker_team") or {}
    composers = ", ".join(c.get("name", "") for c in smt.get("composers", []) if c.get("name"))
    lyricists = ", ".join(l.get("name", "") for l in smt.get("lyricists", []) if l.get("name"))

    # 12. Quality options
    bit_rates = track_info.get("bit_rates") or []
    qualities = _soda_format_qualities(bit_rates)

    # 13. Extra — raw data for further processing
    extra = {
        "tags": [
            t.get("first_level_tag", {}).get("tag_name", "")
            for t in (track_info.get("tags") or [])
            if t.get("first_level_tag")
        ],
        "bit_rates": [
            {"quality": b.get("quality"), "kbps": (b.get("br") or 0) // 1000,
             "size_bytes": b.get("size")}
            for b in bit_rates
        ],
        "vid": audio.get("vid") or "",
        "update_time": audio.get("update_time") or "",
        "group_playable_level": audio.get("group_playable_level") or "",
        "group_download_level": audio.get("group_download_level") or "",
        "sharable_platforms": track_info.get("sharable_platforms") or [],
        "preview_duration_ms": (track_info.get("preview") or {}).get("duration") or 0,
    }

    return SongDetailInfo(
        song_id=song_id,
        platform="soda",
        name=name,
        artist=artist,
        duration=duration,
        cover=cover,
        album=album,
        album_id=album_id,
        publish_date=publish_date,
        favorites=favorites,
        comments=comments,
        shares=shares,
        plays=plays,
        audio_url=audio_url,
        lyrics_lrc=lyrics_lrc,
        genre=genre,
        language=language,
        composers=composers,
        lyricists=lyricists,
        qualities=qualities,
        share_url=share_url,
        resolved_url=final_url,
        extra=extra,
    )

def _scrape_soda_playlist_detail(share_url: str, timeout: int = 15) -> "PlaylistDetailInfo":
    """抓取汽水音乐歌单详情（数据监控，不含歌曲下载）。

    支持:
      - 短链: https://qishui.douyin.com/s/CODE/
      - 完整: https://music.douyin.com/qishui/share/playlist?playlist_id=...

    数据来源: 页面内嵌 _ROUTER_DATA → loaderData → playlist_page
    """
    import datetime

    session = requests.Session()
    session.headers.update({
        "User-Agent": _SCRAPER_MOBILE_UA,
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
    })

    resp = session.get(share_url, timeout=timeout, allow_redirects=True)
    resp.raise_for_status()
    final_url = resp.url
    page_html = resp.text

    m = re.search(
        r'_ROUTER_DATA\s*=\s*(\{.+?\});\s*(?:function|window)',
        page_html,
        re.DOTALL,
    )
    if not m:
        raise ValueError(
            f"无法从页面提取歌单数据 (未找到 _ROUTER_DATA)\nURL: {final_url}"
        )

    router_data = json.loads(m.group(1))
    playlist_page = router_data.get("loaderData", {}).get("playlist_page", {})

    if not playlist_page:
        raise ValueError(
            f"_ROUTER_DATA 中未找到 playlist_page\nURL: {final_url}\n"
            "请确认这是歌单分享链接（而非单曲链接）"
        )

    playlist_info = playlist_page.get("playlistInfo", {})
    medias = playlist_page.get("medias", [])

    # ── 歌单基本信息 ──
    playlist_id = str(playlist_info.get("id") or "")
    title = playlist_info.get("title") or ""

    owner = playlist_info.get("owner") or {}
    creator = owner.get("nickname") or owner.get("name") or ""

    # 封面
    cover = ""
    cover_info = playlist_info.get("cover_url") or {}
    cover_urls = cover_info.get("urls") or []
    cover_uri = cover_info.get("uri") or ""
    cover_tpl = cover_info.get("template_prefix") or ""
    if cover_urls and cover_uri:
        cover = f"{cover_urls[0]}{cover_uri}~{cover_tpl}-image.webp"

    track_count = int(playlist_info.get("count_tracks") or len(medias))
    description = playlist_info.get("description") or ""

    def _ts_to_date(ts) -> str:
        if not ts:
            return ""
        try:
            return datetime.datetime.fromtimestamp(int(ts)).strftime("%Y-%m-%d")
        except Exception:
            return ""

    create_time = _ts_to_date(playlist_info.get("create_time"))
    update_time = _ts_to_date(playlist_info.get("update_time"))

    # ── 提取曲目 ──
    tracks = []
    seen_ids: set = set()
    for media in medias:
        entity = media.get("entity") or {}
        track = entity.get("track") or {}
        if not track.get("id"):
            continue
        tid = str(track["id"])
        if tid in seen_ids:
            continue
        seen_ids.add(tid)
        tracks.append(_soda_track_dict_to_detail(track, share_url=""))

    extra = {
        "play_count": int(playlist_info.get("play_count") or 0),
        "collect_count": int(playlist_info.get("collect_count") or 0),
        "share_count": int(playlist_info.get("share_count") or 0),
    }

    return PlaylistDetailInfo(
        playlist_id=playlist_id,
        platform="soda",
        title=title,
        creator=creator,
        cover=cover,
        track_count=track_count,
        create_time=create_time,
        update_time=update_time,
        description=description,
        tracks=tuple(tracks),
        share_url=share_url,
        resolved_url=final_url,
        extra=extra,
    )


def _scrape_netease_playlist_detail(share_url: str, timeout: int = 15) -> "PlaylistDetailInfo":
    """抓取网易云音乐歌单详情。

    支持:
      - https://music.163.com/playlist?id=xxx
      - https://music.163.com/#/playlist?id=xxx

    数据来源: music.163.com/api/v6/playlist/detail + song/detail + batch comment API
    注意: 网易云公开 API 不提供单曲收藏/分享数（固定 0），但评论数可通过 batch API 获取。
    """
    import datetime
    from urllib.parse import urlparse, parse_qs

    session = requests.Session()
    session.headers.update({
        "User-Agent": _SCRAPER_PC_UA,
        "Referer": "https://music.163.com/",
    })

    # 提取 playlist ID
    resp = session.get(share_url, timeout=timeout, allow_redirects=True)
    resp.raise_for_status()
    final_url = resp.url

    playlist_id = ""
    parsed = urlparse(final_url)
    fragment = parsed.fragment  # /#/playlist?id=xxx
    if fragment:
        fqs = parse_qs(fragment.lstrip("/playlist"))
        playlist_id = (fqs.get("?id") or fqs.get("id") or [""])[0]
    if not playlist_id:
        qs = parse_qs(parsed.query)
        playlist_id = (qs.get("id") or [""])[0]
    if not playlist_id:
        m = re.search(r'/playlist[?/].*?id=(\d+)', final_url)
        if m:
            playlist_id = m.group(1)
    if not playlist_id:
        raise ValueError(f"无法从 URL 提取歌单 ID: {final_url}")

    # 歌单基本信息 + 前N首曲目
    api_resp = session.get(
        f"https://music.163.com/api/v6/playlist/detail?id={playlist_id}&n=1000",
        timeout=timeout,
    )
    data = api_resp.json()
    pl = data.get("playlist", {})
    if not pl:
        raise ValueError(f"网易云 API 未返回歌单数据 (id={playlist_id})")

    # 基本信息
    title = pl.get("name") or ""
    creator = (pl.get("creator") or {}).get("nickname") or ""
    cover = pl.get("coverImgUrl") or ""
    track_count = int(pl.get("trackCount") or 0)
    description = pl.get("description") or ""

    def _ts_to_date(ts_ms) -> str:
        if not ts_ms:
            return ""
        try:
            return datetime.datetime.fromtimestamp(int(ts_ms) / 1000).strftime("%Y-%m-%d")
        except Exception:
            return ""

    create_time = _ts_to_date(pl.get("createTime"))
    update_time = _ts_to_date(pl.get("updateTime"))

    extra = {
        "play_count": int(pl.get("playCount") or 0),
        "collect_count": int(pl.get("subscribedCount") or 0),
        "share_count": int(pl.get("shareCount") or 0),
        "comment_count": int(pl.get("commentCount") or 0),
    }

    # 合并曲目：tracks[] 已经包含大部分，trackIds[] 包含全部
    track_ids_order = [str(t["id"]) for t in pl.get("trackIds", [])]
    tracks_map: dict = {}
    for t in pl.get("tracks", []):
        tracks_map[str(t.get("id"))] = t

    # 补齐缺失曲目
    missing_ids = [tid for tid in track_ids_order if tid not in tracks_map]
    if missing_ids:
        # 批量查询，每次最多 200 个
        for i in range(0, len(missing_ids), 200):
            batch = missing_ids[i:i + 200]
            detail_resp = session.get(
                f"https://music.163.com/api/song/detail/?ids=[{','.join(batch)}]",
                timeout=timeout,
            )
            for s in detail_resp.json().get("songs", []):
                tracks_map[str(s.get("id"))] = s

    # 补齐 publish_date：v6 API 的 publishTime=0 时，改查 song/detail 的 album.publishTime
    publish_ts_overrides: dict = {}
    no_date_ids = [
        tid for tid in tracks_map
        if not int(tracks_map[tid].get("publishTime") or 0)
        and not int(
            (tracks_map[tid].get("al") or tracks_map[tid].get("album") or {}).get("publishTime") or 0
        )
    ]
    if no_date_ids:
        for i in range(0, len(no_date_ids), 200):
            batch = no_date_ids[i:i + 200]
            try:
                detail_resp = session.get(
                    f"https://music.163.com/api/song/detail/?ids=[{','.join(batch)}]",
                    timeout=timeout,
                )
                for s in detail_resp.json().get("songs", []):
                    sid = str(s.get("id"))
                    pts = (s.get("album") or {}).get("publishTime")
                    if pts:
                        publish_ts_overrides[sid] = int(pts)
            except Exception:
                pass

    def _netease_track_to_detail(t: dict, comments: int = 0) -> "SongDetailInfo":
        song_id = str(t.get("id") or "")
        name = t.get("name") or ""
        # v6 API uses 'ar', older uses 'artists'
        artists = t.get("ar") or t.get("artists") or []
        artist = "、".join(a.get("name", "") for a in artists if a.get("name"))
        duration = int((t.get("dt") or t.get("duration") or 0) // 1000)
        al = t.get("al") or t.get("album") or {}
        album = al.get("name") or ""
        album_id = str(al.get("id") or "")
        cover_url = al.get("picUrl") or ""

        publish_ts = (
            t.get("publishTime")
            or al.get("publishTime")
            or publish_ts_overrides.get(song_id)
        )
        publish_date = _ts_to_date(publish_ts) if publish_ts else ""

        resolved_url = f"https://music.163.com/song?id={song_id}" if song_id else ""

        return SongDetailInfo(
            song_id=song_id,
            platform="netease",
            name=name,
            artist=artist,
            duration=duration,
            cover=cover_url,
            album=album,
            album_id=album_id,
            publish_date=publish_date,
            favorites=0,
            comments=comments,
            shares=0,
            plays=0,
            audio_url="",
            lyrics_lrc="",
            genre="",
            language="",
            composers="",
            lyricists="",
            qualities="",
            share_url="",
            resolved_url=resolved_url,
            extra={},
        )

    # 批量获取评论数（/api/batch 一次请求可查多首）
    comment_counts: dict = {}
    all_ids = list(tracks_map.keys())
    # batch API 每次最多约 100 个子请求
    for i in range(0, len(all_ids), 100):
        batch_ids = all_ids[i:i + 100]
        batch_payload = {
            f"/api/v1/resource/comments/R_SO_4_{sid}": '{"limit":0}'
            for sid in batch_ids
        }
        try:
            batch_resp = session.post(
                "https://music.163.com/api/batch",
                data=batch_payload,
                timeout=timeout,
            )
            batch_data = batch_resp.json()
            for sid in batch_ids:
                key = f"/api/v1/resource/comments/R_SO_4_{sid}"
                sub = batch_data.get(key, {})
                comment_counts[sid] = int(sub.get("total") or 0)
        except Exception:
            pass  # 降级：评论数为 0

    tracks = []
    for tid in track_ids_order:
        t = tracks_map.get(tid)
        if t:
            tracks.append(_netease_track_to_detail(t, comments=comment_counts.get(tid, 0)))

    return PlaylistDetailInfo(
        playlist_id=playlist_id,
        platform="netease",
        title=title,
        creator=creator,
        cover=cover,
        track_count=track_count,
        create_time=create_time,
        update_time=update_time,
        description=description,
        tracks=tuple(tracks),
        share_url=share_url,
        resolved_url=final_url,
        extra=extra,
    )


def _scrape_netease_detail(share_url: str, timeout: int = 15) -> SongDetailInfo:
    """抓取网易云音乐歌曲详情。

    支持:
      - https://music.163.com/#/song?id=xxx
      - https://music.163.com/song?id=xxx
      - https://163.cn/xxx (短链)
    """
    import datetime
    from urllib.parse import urlparse, parse_qs

    session = requests.Session()
    session.headers.update({
        "User-Agent": _SCRAPER_PC_UA,
        "Referer": "https://music.163.com/",
    })

    # Follow redirect
    resp = session.get(share_url, timeout=timeout, allow_redirects=True)
    final_url = resp.url

    # Extract song_id — handle hash-based URL (#/song?id=xxx)
    song_id = ""
    parsed = urlparse(final_url)

    fragment = parsed.fragment
    if fragment:
        # fragment: /song?id=xxx
        fqs = parse_qs(fragment.lstrip("/song"))
        song_id = (fqs.get("?id") or fqs.get("id") or [""])[0]

    if not song_id:
        qs = parse_qs(parsed.query)
        song_id = (qs.get("id") or [""])[0]

    if not song_id:
        m = re.search(r'/song(?:Detail)?[/=](\d+)', final_url)
        if m:
            song_id = m.group(1)

    if not song_id:
        raise ValueError(f"无法从 URL 提取歌曲 ID: {final_url}")

    # Netease API
    api_url = f"https://music.163.com/api/song/detail/?id={song_id}&ids=[{song_id}]"
    api_resp = session.get(api_url, timeout=10)
    data = api_resp.json()
    songs = data.get("songs", [])
    if not songs:
        raise ValueError(f"网易云 API 未返回数据 (id={song_id})")

    track = songs[0]
    name = track.get("name") or ""

    artists = track.get("artists") or []
    artist = "、".join(a.get("name", "") for a in artists if a.get("name"))

    duration = int((track.get("duration") or 0) // 1000)  # ms → s

    album_info = track.get("album") or {}
    album = album_info.get("name") or ""
    cover = album_info.get("picUrl") or ""

    publish_date = ""
    publish_time = album_info.get("publishTime")
    if publish_time:
        publish_date = datetime.datetime.fromtimestamp(publish_time / 1000).strftime("%Y-%m-%d")

    return SongDetailInfo(
        song_id=song_id,
        platform="netease",
        name=name,
        artist=artist,
        duration=duration,
        cover=cover,
        album=album,
        album_id=str(album_info.get("id") or ""),
        publish_date=publish_date,
        favorites=0,
        comments=0,
        shares=0,
        plays=0,
        audio_url="",
        lyrics_lrc="",
        genre="",
        language="",
        composers="",
        lyricists="",
        qualities="",
        share_url=share_url,
        resolved_url=final_url,
        extra=track,
    )


def _kugou_search_url(name: str, artist: str, song_hash: str = "") -> str:
    """构造酷狗歌曲链接。

    优先使用 hash 生成分享链接，如果没有 hash 则使用搜索链接。
    """
    if song_hash:
        # 使用 hash 生成分享链接（可直接播放）
        return f"https://www.kugou.com/song/#hash={song_hash}"

    # 回退到搜索链接
    from urllib.parse import quote_plus
    kw = f"{name} {artist}".strip()
    return (
        f"https://www.kugou.com/search.html#searchType=song"
        f"&searchKeyWord={quote_plus(kw)}"
    ) if kw else ""


def _scrape_kugou_song_detail(share_url: str, timeout: int = 15) -> "SongDetailInfo":
    """抓取酷狗音乐单曲详情。

    支持:
      - https://m.kugou.com/share/song.html?chain=xxx
      - https://www.kugou.com/share/xxx.html

    数据来源: 页面内嵌 var phpParam = {...}
    """
    import json

    session = requests.Session()
    session.headers.update({
        "User-Agent": (
            "Mozilla/5.0 (iPhone; CPU iPhone OS 13_2_3 like Mac OS X) "
            "AppleWebKit/605.1.15 (KHTML, like Gecko) "
            "Version/13.0.3 Mobile/15E148 Safari/604.1"
        ),
        "Referer": "http://m.kugou.com",
    })

    resp = session.get(share_url, timeout=timeout, allow_redirects=True)
    resp.raise_for_status()
    html = resp.text

    # 查找 phpParam 变量
    match = re.search(r'var\s+phpParam\s*=\s*(\{[^;]+\});', html)
    if not match:
        raise ValueError(f"无法从酷狗单曲页面提取数据 (未找到 phpParam)\nURL: {resp.url}")

    data = json.loads(match.group(1))
    song_info = data.get("song_info", {}).get("data", {})

    # 解析歌曲信息
    filename = song_info.get("fileName", "")
    if " - " in filename:
        artist_part, name_part = filename.split(" - ", 1)
    else:
        artist_part, name_part = "", filename

    song_hash = data.get("hash", "")
    duration = int(data.get("timelen", 0)) // 1000
    album_id = str(data.get("album_id", ""))

    # 歌手信息
    authors = song_info.get("authors", [])
    if authors and not artist_part:
        artist_part = authors[0].get("author_name", "")

    cover = ""
    if authors:
        cover = authors[0].get("avatar", "").replace("{size}", "240")

    return SongDetailInfo(
        song_id=song_hash,
        platform="kugou",
        name=name_part,
        artist=artist_part,
        duration=duration,
        cover=cover,
        album="",
        album_id=album_id,
        publish_date="",
        favorites=0,
        comments=0,
        shares=0,
        plays=0,
        audio_url="",
        lyrics_lrc="",
        genre="",
        language="",
        composers="",
        lyricists="",
        qualities="",
        share_url=share_url,
        resolved_url=_kugou_search_url(name_part, artist_part, song_hash),
        extra=data,
    )


def _fetch_kugou_playlist_via_api(gcid: str, timeout: int = 15) -> "PlaylistDetailInfo":
    """通过 KuGouMusicApi 服务获取酷狗歌单完整信息。

    Args:
        gcid: 歌单 global_collection_id (格式: collection_3_xxx_x_x)
        timeout: 请求超时时间

    Returns:
        PlaylistDetailInfo 对象

    Raises:
        requests.RequestException: API 服务不可用或请求失败
        ValueError: 返回数据格式错误
    """
    api_base = os.environ.get("KUGOU_API_URL", DEFAULT_KUGOU_API_URL)

    # 1. 获取歌单详情
    detail_url = f"{api_base}/playlist/detail"
    detail_resp = requests.get(detail_url, params={"ids": gcid}, timeout=timeout)
    detail_resp.raise_for_status()
    detail_data = detail_resp.json()

    if not detail_data.get("data"):
        raise ValueError(f"KuGouMusicApi 返回错误: {detail_data}")

    playlist_info = detail_data["data"][0] if isinstance(detail_data["data"], list) else detail_data["data"]

    # 检查返回的 code 字段（在 playlist_info 内部）
    if playlist_info.get("code") != 1:
        raise ValueError(f"KuGouMusicApi 返回错误: {playlist_info}")

    # 2. 获取所有歌曲（分页）
    all_tracks = []
    page = 1
    pagesize = 100

    while True:
        tracks_url = f"{api_base}/playlist/track/all"
        tracks_resp = requests.get(
            tracks_url,
            params={"id": gcid, "page": page, "pagesize": pagesize},
            timeout=timeout
        )
        tracks_resp.raise_for_status()
        tracks_data = tracks_resp.json()

        if tracks_data.get("code") != 200:
            break

        songs = tracks_data.get("data", {}).get("info", [])
        if not songs:
            break

        for song in songs:
            # 解析歌曲信息
            filename = song.get("filename", "")
            if " - " in filename:
                artist_part, name_part = filename.split(" - ", 1)
            else:
                artist_part, name_part = "", filename

            song_hash = song.get("hash", "")
            duration = int(song.get("duration", 0))
            album_id = str(song.get("album_id", ""))

            all_tracks.append(SongDetailInfo(
                song_id=song_hash,
                platform="kugou",
                name=name_part,
                artist=artist_part,
                duration=duration,
                cover="",
                album="",
                album_id=album_id,
                publish_date="",
                favorites=0,
                comments=0,
                shares=0,
                plays=0,
                audio_url="",
                lyrics_lrc="",
                genre="",
                language="",
                composers="",
                lyricists="",
                qualities="",
                share_url="",
                resolved_url=_kugou_search_url(name_part, artist_part, song_hash),
                extra=song,
            ))

        # 检查是否还有更多页
        total = tracks_data.get("data", {}).get("total", 0)
        if len(all_tracks) >= total:
            break
        page += 1

    # 3. 构造返回对象
    return PlaylistDetailInfo(
        playlist_id=gcid,
        platform="kugou",
        title=playlist_info.get("name", ""),
        creator=playlist_info.get("list_create_username", ""),
        cover=playlist_info.get("pic", "").replace("{size}", "400"),
        track_count=len(all_tracks),
        create_time="",
        update_time="",
        description=playlist_info.get("intro", ""),
        tracks=tuple(all_tracks),
        share_url=f"https://www.kugou.com/songlist/{gcid}/",
        resolved_url=f"https://www.kugou.com/songlist/{gcid}/",
        extra={
            "play_count": 0,
            "collect_count": playlist_info.get("collect_total", 0),
            "share_count": 0,
            "comment_count": 0,
        },
    )


def _scrape_kugou_zlist(zlist_url: str, share_url: str = "", original_url: str = "",
                        timeout: int = 15) -> "PlaylistDetailInfo":
    """解析酷狗 wwwapi.kugou.com/share/zlist.html 格式的歌单分享页面。

    支持两种格式：
      Format 1（有 global_collection_id，~669KB）：歌曲在 dataobj='[{...}]' HTML 属性中
      Format 2（无 global_collection_id，~37KB，带 listid）：歌曲在 var dataFromSmarty = [...] 中
        字段: hash / timelength / author_name / song_name / album_id
    """
    import html as _html
    from urllib.parse import urlparse, parse_qs

    session = requests.Session()
    session.headers.update({
        "User-Agent": _SCRAPER_PC_UA,
        "Referer": "https://www.kugou.com/",
    })

    resp = session.get(zlist_url, timeout=timeout, allow_redirects=True)
    resp.raise_for_status()
    html_text = resp.text

    # 从 URL 参数提取 uid / global_collection_id / chain
    parsed_qs = parse_qs(urlparse(zlist_url).query)
    uid = (parsed_qs.get("uid") or [""])[0]
    gcid_raw = (parsed_qs.get("global_collection_id") or [""])[0]
    chain_code = (parsed_qs.get("chain") or [""])[0]

    def _make_title(text: str) -> str:
        t = re.sub(r'[_|]?(酷狗音乐|kugou).*$', '', text, flags=re.I).strip()
        t = re.sub(r'^酷狗.*$', '', t, flags=re.I).strip()
        return _html.unescape(t)

    title_m = re.search(r'<title>([^<]+)</title>', html_text)
    raw_title = _make_title(title_m.group(1) if title_m else "")
    chain_hint = f"酷狗歌单 ({chain_code})" if chain_code else "酷狗歌单"
    playlist_title = raw_title or gcid_raw or chain_hint

    tracks = []

    # ── 格式路由：有 global_collection_id → Format 1；否则 → Format 2 ──────────
    use_format1 = bool(gcid_raw)

    # ── Format 1: dataobj 属性（带 global_collection_id 的大页面）──────────────
    if use_format1:
        dataobj_list = re.findall(r"dataobj='([^']+)'", html_text)
        seen_hashes: set = set()
        for raw_obj in dataobj_list:
            try:
                decoded = _html.unescape(raw_obj)
                arr = json.loads(decoded)
                if not isinstance(arr, list) or not arr:
                    continue
                s = arr[0]
                song_hash = (s.get("Hash") or s.get("hash") or "").upper()
                if not song_hash or song_hash in seen_hashes:
                    continue
                seen_hashes.add(song_hash)
                duration = int((s.get("timeLen") or s.get("timelength") or 0) // 1000)
                file_name = _html.unescape(s.get("FileName") or s.get("audio_name") or "")
                if " - " in file_name:
                    artist_part, name_part = file_name.split(" - ", 1)
                else:
                    artist_part, name_part = "", file_name
                tracks.append(SongDetailInfo(
                    song_id=song_hash, platform="kugou",
                    name=name_part, artist=artist_part,
                    duration=duration, cover="", album="",
                    album_id=str(s.get("albumId") or ""),
                    publish_date="", favorites=0, comments=0,
                    shares=0, plays=0, audio_url="", lyrics_lrc="",
                    genre="", language="", composers="", lyricists="",
                    qualities="", share_url=share_url,
                    resolved_url=_kugou_search_url(name_part, artist_part, song_hash),
                    extra={},
                ))
            except Exception:
                pass

    # ── Format 2: dataFromSmarty（无 global_collection_id，需保留 listid）────────
    if not use_format1:
        # 必须用原始 URL（含 listid）请求，listid-stripped URL 会返回首页
        target_html = html_text
        if original_url and original_url != zlist_url:
            orig_resp = session.get(original_url.split('#')[0], timeout=timeout)
            target_html = orig_resp.text
            # 从单曲页标题尝试提取歌单标题（歌曲名 - 歌手名）
            title_m2 = re.search(r'<title>([^<]+)</title>', target_html)
            if title_m2:
                t2 = _make_title(title_m2.group(1))
                if t2:
                    playlist_title = t2

        smarty_m = re.search(r'var dataFromSmarty\s*=\s*(\[.+)', target_html)
        if smarty_m:
            raw_arr = smarty_m.group(1).rstrip().rstrip(';')
            decoder = json.JSONDecoder()
            raw_inner = raw_arr[1:]  # skip leading [
            while raw_inner.strip():
                raw_inner = raw_inner.strip().lstrip(',')
                if raw_inner.startswith(']'):
                    break
                try:
                    s, end = decoder.raw_decode(raw_inner)
                    raw_inner = raw_inner[end:]
                except Exception:
                    break
                song_hash = (s.get("hash") or "").upper()
                if not song_hash:
                    continue
                duration = int((s.get("timelength") or 0) // 1000)
                artist_part = _html.unescape(s.get("author_name") or "")
                name_part = _html.unescape(s.get("song_name") or s.get("audio_name") or "")
                # audio_name fallback: "Artist - SongName"
                if not name_part and s.get("audio_name") and " - " in s["audio_name"]:
                    artist_part, name_part = s["audio_name"].split(" - ", 1)
                tracks.append(SongDetailInfo(
                    song_id=song_hash, platform="kugou",
                    name=_html.unescape(name_part),
                    artist=artist_part,
                    duration=duration, cover="", album="",
                    album_id=str(s.get("album_id") or ""),
                    publish_date="", favorites=0, comments=0,
                    shares=0, plays=0, audio_url="", lyrics_lrc="",
                    genre="", language="", composers="", lyricists="",
                    qualities="", share_url=share_url,
                    resolved_url=_kugou_search_url(
                        _html.unescape(name_part), artist_part, song_hash
                    ),
                    extra={},
                ))

    return PlaylistDetailInfo(
        playlist_id=gcid_raw or uid or "unknown",
        platform="kugou",
        title=playlist_title,
        creator=uid or "",
        cover="",
        track_count=len(tracks),
        create_time="",
        update_time="",
        description="",
        tracks=tuple(tracks),
        share_url=share_url,
        resolved_url=zlist_url,
        extra={},
    )


def _scrape_kugou_playlist_detail(share_url: str, timeout: int = 15) -> "PlaylistDetailInfo":
    """抓取酷狗音乐歌单详情。

    支持:
      - https://www.kugou.com/songlist/gcid_xxx/
      - https://m.kugou.com/songlist/gcid_xxx/

    策略:
      1. 优先使用 KuGouMusicApi 服务（如果 KUGOU_API_URL 可用）
      2. 回退到网页抓取（SSR 限制，最多返回 10 首）

    注意:
      - KuGouMusicApi 可获取完整歌单，无 SSR 限制
      - 网页抓取需要随机 X-Forwarded-For 绕过限速
      - 酷狗公开 API 不提供单曲评论/收藏/分享数，这些字段固定为 0
    """
    import datetime
    import random
    import html as _html

    # 直接传入 zlist.html URL（去掉 listid 和 fragment 后直接解析）
    if "zlist.html" in share_url or "wwwapi.kugou.com/share" in share_url:
        orig = share_url.split('#')[0]
        clean = re.sub(r'[&?]listid=\d+', '', orig)
        return _scrape_kugou_zlist(clean, share_url=share_url, original_url=orig, timeout=timeout)

    # 短链 (t1.kugou.com/xxx) 先跟重定向，再判断格式
    if re.match(r'https?://t\d+\.kugou\.com/', share_url):
        _redir = requests.Session()
        _redir.headers["User-Agent"] = _SCRAPER_PC_UA
        _redir_resp = _redir.get(share_url, timeout=timeout, allow_redirects=True)
        _redir_url = _redir_resp.url

        if "zlist.html" in _redir_url or "wwwapi.kugou.com/share" in _redir_url:
            _zlist_url = re.sub(r'[&?]listid=\d+', '', _redir_url)
            return _scrape_kugou_zlist(
                _zlist_url, share_url=share_url,
                original_url=_redir_url, timeout=timeout,
            )

        # 重定向后尝试再提取 gcid
        share_url = _redir_url

    gcid_match = re.search(r'(gcid_[a-zA-Z0-9]+)', share_url)
    if not gcid_match:
        raise ValueError(f"无法从 URL 提取酷狗歌单 ID (gcid_xxx): {share_url}")
    gcid = gcid_match.group(1)

    # 尝试使用 KuGouMusicApi（优先）
    # 注意：gcid 格式需要先从网页获取完整的 collection_id
    # 因为 gcid_xxx 无法直接转换为 collection_3_xxx_x_x（缺少 listid）
    # 所以这里先尝试网页抓取获取 collection_id，如果失败再用 API
    api_base = os.environ.get("KUGOU_API_URL", DEFAULT_KUGOU_API_URL)

    # 先尝试从网页获取 collection_id
    try:
        rand_ip = (
            f"{random.randint(1, 254)}.{random.randint(0, 254)}"
            f".{random.randint(0, 254)}.{random.randint(1, 254)}"
        )
        session = requests.Session()
        session.headers.update({
            "User-Agent": (
                "Mozilla/5.0 (iPhone; CPU iPhone OS 13_2_3 like Mac OS X) "
                "AppleWebKit/605.1.15 (KHTML, like Gecko) "
                "Version/13.0.3 Mobile/15E148 Safari/604.1"
            ),
            "Referer": "http://m.kugou.com",
            "X-Forwarded-For": rand_ip,
            "X-Real-IP": rand_ip,
        })

        resp = session.get(
            f"https://www.kugou.com/songlist/{gcid}/",
            timeout=5,
            allow_redirects=True,
        )
        resp.raise_for_status()
        html = resp.text

        # 尝试从页面提取 collection_id
        collection_match = re.search(r'(collection_\d+_\d+_\d+_\d+)', html)
        if collection_match:
            collection_id = collection_match.group(1)
            # 尝试使用 API
            try:
                test_resp = requests.get(
                    f"{api_base}/playlist/detail",
                    params={"ids": collection_id},
                    timeout=1
                )
                if test_resp.status_code == 200:
                    api_data = test_resp.json()
                    if api_data.get("data") and len(api_data["data"]) > 0:
                        playlist_info = api_data["data"][0]
                        if playlist_info.get("code") == 1:
                            return _fetch_kugou_playlist_via_api(collection_id, timeout=timeout)
            except:
                pass
    except:
        pass

    # 回退：网页抓取（SSR 限制，最多 10 首）
    rand_ip = (
        f"{random.randint(1, 254)}.{random.randint(0, 254)}"
        f".{random.randint(0, 254)}.{random.randint(1, 254)}"
    )
    session = requests.Session()
    session.headers.update({
        "User-Agent": (
            "Mozilla/5.0 (iPhone; CPU iPhone OS 13_2_3 like Mac OS X) "
            "AppleWebKit/605.1.15 (KHTML, like Gecko) "
            "Version/13.0.3 Mobile/15E148 Safari/604.1"
        ),
        "Referer": "http://m.kugou.com",
        "X-Forwarded-For": rand_ip,
        "X-Real-IP": rand_ip,
    })

    resp = session.get(
        f"https://www.kugou.com/songlist/{gcid}/",
        timeout=timeout,
        allow_redirects=True,
    )
    resp.raise_for_status()
    final_url = resp.url
    html = resp.text

    m = re.search(
        r"window\.\$output\s*=\s*(\{.+?\})\s*;\s*</script>",
        html,
        re.DOTALL,
    )
    if not m:
        raise ValueError(
            f"无法从酷狗歌单页面提取数据 (未找到 window.$output)\n"
            f"URL: {final_url}\n"
            "提示: 酷狗歌单抓取需要随机 IP header 绕过限速，请重试。"
        )

    data = json.loads(m.group(1))
    listinfo = data.get("info", {}).get("listinfo", {})
    songs_raw = data.get("info", {}).get("songs", [])

    # 歌单基本信息
    title = listinfo.get("name") or ""
    creator = listinfo.get("list_create_username") or ""
    pic = listinfo.get("pic") or ""
    cover = pic.replace("{size}", "240") if pic else ""
    track_count = int(listinfo.get("count") or 0)
    description = listinfo.get("intro") or ""

    extra = {
        "play_count": int(listinfo.get("heat") or 0),
        "collect_count": int(listinfo.get("collect_count") or 0),
        "share_count": 0,
        "comment_count": int(listinfo.get("comment_count") or 0),
    }

    # 解析曲目（SSR 最多返回 10 首）
    def _kugou_track_to_detail(t: dict) -> "SongDetailInfo":
        name = t.get("name") or ""
        singers = t.get("singerinfo") or []
        artist = "、".join(s.get("name", "") for s in singers if s.get("name"))
        duration = int((t.get("timelen") or 0) // 1000)
        al_info = t.get("albuminfo") or {}
        album = al_info.get("name") or ""
        album_id = str(t.get("album_id") or "")
        song_hash = t.get("hash") or ""
        raw_cover = t.get("cover") or ""
        cover_url = raw_cover.replace("{size}", "240") if raw_cover else ""
        resolved_url = _kugou_search_url(name, artist, song_hash) if (name or artist or song_hash) else ""

        return SongDetailInfo(
            song_id=song_hash,
            platform="kugou",
            name=name,
            artist=artist,
            duration=duration,
            cover=cover_url,
            album=album,
            album_id=album_id,
            publish_date="",
            favorites=0,
            comments=0,
            shares=0,
            plays=0,
            audio_url="",
            lyrics_lrc="",
            genre="",
            language="",
            composers="",
            lyricists="",
            qualities="",
            share_url="",
            resolved_url=resolved_url,
            extra={},
        )

    tracks = [_kugou_track_to_detail(t) for t in songs_raw]

    return PlaylistDetailInfo(
        playlist_id=gcid,
        platform="kugou",
        title=title,
        creator=creator,
        cover=cover,
        track_count=track_count,
        create_time="",
        update_time="",
        description=description,
        tracks=tuple(tracks),
        share_url=share_url,
        resolved_url=final_url,
        extra=extra,
    )


def _scrape_qq_playlist_detail(share_url: str, timeout: int = 15) -> "PlaylistDetailInfo":
    """抓取 QQ 音乐歌单详情。

    支持:
      - https://y.qq.com/n/ryqq_v2/playlist/DISSTID
      - https://y.qq.com/n/ryqq/playlist/DISSTID

    数据来源: c.y.qq.com/qzone/fcg-bin/fcg_ucc_getcdinfo_byids_cp.fcg
    注意: QQ 音乐公开 API 不提供单曲评论/收藏/分享数，这些字段固定为 0。
    """
    import datetime
    from urllib.parse import urlparse, parse_qs

    session = requests.Session()
    session.headers.update({
        "User-Agent": _SCRAPER_PC_UA,
        "Referer": "https://y.qq.com/",
    })

    # 提取 playlist ID (disstid)
    resp = session.get(share_url, timeout=timeout, allow_redirects=True)
    final_url = resp.url

    disstid = ""
    m = re.search(r'/playlist/(\d+)', final_url)
    if m:
        disstid = m.group(1)
    if not disstid:
        qs = parse_qs(urlparse(final_url).query)
        disstid = (qs.get("id") or qs.get("disstid") or [""])[0]
    if not disstid:
        raise ValueError(f"无法从 URL 提取歌单 ID: {final_url}")

    # 旧版歌单 API（稳定可用，无需签名）
    api_url = (
        f"https://c.y.qq.com/qzone/fcg-bin/fcg_ucc_getcdinfo_byids_cp.fcg"
        f"?type=1&json=1&utf8=1&onlysong=0&new_format=1&disstid={disstid}&format=json"
    )
    api_resp = session.get(api_url, timeout=timeout)
    data = api_resp.json()
    cdlist = data.get("cdlist", [])
    if not cdlist:
        raise ValueError(f"QQ 音乐 API 未返回歌单数据 (disstid={disstid})")

    cd = cdlist[0]
    title = cd.get("dissname") or ""
    creator = cd.get("nickname") or cd.get("nick") or ""
    cover = cd.get("logo") or ""
    track_count = int(cd.get("total_song_num") or cd.get("songnum") or 0)
    description = cd.get("desc") or ""

    def _ts_to_date(ts) -> str:
        if not ts:
            return ""
        try:
            return datetime.datetime.fromtimestamp(int(ts)).strftime("%Y-%m-%d")
        except Exception:
            return ""

    create_time = _ts_to_date(cd.get("ctime"))
    update_time = _ts_to_date(cd.get("mtime") or cd.get("song_update_time"))

    extra = {
        "play_count": int(cd.get("visitnum") or 0),
        "collect_count": 0,
        "share_count": 0,
        "comment_count": int(cd.get("cmtnum") or 0),
    }

    # 解析曲目列表
    def _qq_track_to_detail(t: dict) -> "SongDetailInfo":
        song_mid = str(t.get("mid") or "")
        song_id = str(t.get("id") or "")
        name = t.get("name") or ""
        singers = t.get("singer") or []
        artist = "、".join(s.get("name", "") for s in singers if s.get("name"))
        duration = int(t.get("interval") or 0)
        al = t.get("album") or {}
        album = al.get("name") or ""
        album_id = str(al.get("mid") or al.get("id") or "")
        pmid = al.get("pmid") or ""
        cover_url = (
            f"https://y.gtimg.cn/music/photo_new/T002R300x300M000{pmid}.jpg"
            if pmid else ""
        )
        publish_date = t.get("time_public") or ""
        resolved_url = (
            f"https://y.qq.com/n/ryqq/songDetail/{song_mid}" if song_mid else ""
        )

        return SongDetailInfo(
            song_id=song_mid or song_id,
            platform="qq",
            name=name,
            artist=artist,
            duration=duration,
            cover=cover_url,
            album=album,
            album_id=album_id,
            publish_date=publish_date,
            favorites=0,
            comments=0,
            shares=0,
            plays=0,
            audio_url="",
            lyrics_lrc="",
            genre="",
            language="",
            composers="",
            lyricists="",
            qualities="",
            share_url="",
            resolved_url=resolved_url,
            extra={},
        )

    tracks = [_qq_track_to_detail(t) for t in cd.get("songlist", [])]

    return PlaylistDetailInfo(
        playlist_id=disstid,
        platform="qq",
        title=title,
        creator=creator,
        cover=cover,
        track_count=track_count,
        create_time=create_time,
        update_time=update_time,
        description=description,
        tracks=tuple(tracks),
        share_url=share_url,
        resolved_url=final_url,
        extra=extra,
    )


def _scrape_qq_detail(share_url: str, timeout: int = 15) -> SongDetailInfo:
    """抓取 QQ 音乐歌曲详情。

    支持:
      - https://y.qq.com/n/ryqq/songDetail/SONGMID
      - https://c6.y.qq.com/...
    """
    from urllib.parse import urlparse, parse_qs, quote

    session = requests.Session()
    session.headers.update({
        "User-Agent": _SCRAPER_PC_UA,
        "Referer": "https://y.qq.com/",
    })

    # Follow redirect
    resp = session.get(share_url, timeout=timeout, allow_redirects=True)
    final_url = resp.url

    # Extract songmid
    songmid = ""
    m = re.search(r'/songDetail/([A-Za-z0-9]+)', final_url)
    if m:
        songmid = m.group(1)
    if not songmid:
        qs = parse_qs(urlparse(final_url).query)
        songmid = (qs.get("songmid") or qs.get("mid") or [""])[0]
    if not songmid:
        raise ValueError(f"无法从 URL 提取 songmid: {final_url}")

    # QQ Music API
    req_data = {
        "comm": {"ct": 24, "cv": 0},
        "detail": {
            "module": "music.pf_song_detail_svr",
            "method": "get_song_detail_yqq",
            "param": {"song_mid": songmid},
        },
    }
    api_url = (
        "https://u.y.qq.com/cgi-bin/musics.fcg?format=json&data="
        + quote(json.dumps(req_data, separators=(",", ":")))
    )
    track: dict = {}
    try:
        api_resp = session.get(api_url, timeout=10)
        data = api_resp.json()
        track = (
            data.get("detail", {})
            .get("data", {})
            .get("track_info", {})
        )
    except Exception:
        pass

    name = track.get("name") or ""
    singers = track.get("singer") or []
    artist = "、".join(s.get("name", "") for s in singers if s.get("name"))
    duration = int(track.get("interval") or 0)

    album_info = track.get("album") or {}
    album = album_info.get("name") or ""
    pmid = album_info.get("pmid") or ""
    cover = f"https://y.gtimg.cn/music/photo_new/T002R800x800M000{pmid}.jpg" if pmid else ""

    return SongDetailInfo(
        song_id=songmid,
        platform="qq",
        name=name,
        artist=artist,
        duration=duration,
        cover=cover,
        album=album,
        album_id=str(album_info.get("mid") or album_info.get("id") or ""),
        publish_date="",
        favorites=0,
        comments=0,
        shares=0,
        plays=0,
        audio_url="",
        lyrics_lrc="",
        genre="",
        language="",
        composers="",
        lyricists="",
        qualities="",
        share_url=share_url,
        resolved_url=final_url,
        extra=track,
    )


def _soda_track_dict_to_detail(track: dict, share_url: str = "") -> "SongDetailInfo":
    """将汽水音乐 trackInfo dict（来自 artistTracks/relatedTracks 等）转为 SongDetailInfo。

    这类 track dict 已包含 id/name/artists/duration/album/stats，
    可直接使用无需额外网络请求。但没有 audio_url 和 lyrics，需要单独抓取。
    """
    import datetime

    song_id = str(track.get("id") or "")
    name = track.get("name") or ""

    artists_list = track.get("artists") or []
    artist = "\u3001".join(a.get("name", "") for a in artists_list if a.get("name"))

    duration = int((track.get("duration") or 0) / 1000)  # ms → s

    album_info = track.get("album") or {}
    album = album_info.get("name") or ""
    album_id = str(album_info.get("id") or "")

    # Cover from album url_cover
    cover = ""
    url_cover = album_info.get("url_cover") or {}
    url_bases = url_cover.get("urls") or []
    uri = url_cover.get("uri") or ""
    tpl = url_cover.get("template_prefix") or ""
    if url_bases and uri:
        cover = f"{url_bases[0]}{uri}~{tpl}-image.webp"

    # Stats
    stats = track.get("stats") or {}
    favorites = int(stats.get("count_collected") or 0)
    comments = int(stats.get("count_comment") or 0)
    shares = int(stats.get("count_shared") or 0)
    plays = int(stats.get("count_played") or stats.get("count_view") or 0)

    # Publish date
    release_ts = int(album_info.get("release_date") or 0)
    publish_date = ""
    if release_ts:
        publish_date = datetime.datetime.fromtimestamp(release_ts).strftime("%Y-%m-%d")

    # Composer/lyricist
    smt = track.get("song_maker_team") or {}
    composers = ", ".join(c.get("name", "") for c in smt.get("composers", []) if c.get("name"))
    lyricists = ", ".join(l.get("name", "") for l in smt.get("lyricists", []) if l.get("name"))

    bit_rates = track.get("bit_rates") or []
    qualities = _soda_format_qualities(bit_rates)

    resolved_url = (
        f"https://music.douyin.com/qishui/share/track?track_id={song_id}"
        if song_id else ""
    )

    return SongDetailInfo(
        song_id=song_id,
        platform="soda",
        name=name,
        artist=artist,
        duration=duration,
        cover=cover,
        album=album,
        album_id=album_id,
        publish_date=publish_date,
        favorites=favorites,
        comments=comments,
        shares=shares,
        plays=plays,
        audio_url="",
        lyrics_lrc="",
        genre="",
        language="",
        composers=composers,
        lyricists=lyricists,
        qualities=qualities,
        share_url=share_url,
        resolved_url=resolved_url,
        extra={"bit_rates": [{"quality": b.get("quality"), "kbps": (b.get("br") or 0) // 1000} for b in bit_rates]},
    )


def get_soda_related_tracks(share_url: str, timeout: int = 15) -> list:
    """从单曲分享页中提取同艺人曲目和相关曲目（无需额外请求，已含统计数据）。

    返回 list[SongDetailInfo]，每条包含完整统计数据但无 audio_url/lyrics。
    """
    session = requests.Session()
    session.headers.update({
        "User-Agent": _SCRAPER_MOBILE_UA,
        "Accept-Language": "zh-CN,zh;q=0.9",
    })

    resp = session.get(share_url, timeout=timeout, allow_redirects=True)
    resp.raise_for_status()
    page_html = resp.text

    m = re.search(
        r'_ROUTER_DATA\s*=\s*(\{.+?\});\s*(?:function|window)',
        page_html,
        re.DOTALL,
    )
    if not m:
        return []

    router_data = json.loads(m.group(1))
    audio = (
        router_data.get("loaderData", {})
        .get("track_page", {})
        .get("audioWithLyricsOption", {})
    )

    results = []
    seen_ids: set = set()

    def _add_tracks(track_list: list) -> None:
        for item in track_list:
            track = item.get("track") or item  # artistTracks wrap in {track: {...}}
            tid = str(track.get("id") or "")
            if tid and tid not in seen_ids:
                seen_ids.add(tid)
                results.append(_soda_track_dict_to_detail(track))

    _add_tracks(audio.get("artistTracks") or [])
    _add_tracks(audio.get("relatedTracks") or [])

    return results


def get_song_detail_from_url(url: str, timeout: int = 15) -> SongDetailInfo:
    """从分享链接抓取歌曲详情（收藏、评论、分享等）。

    支持平台:
      - 汽水音乐 (qishui.douyin.com) — 含完整统计数据
      - 网易云音乐 (music.163.com)
      - QQ 音乐 (y.qq.com)
      - 酷狗音乐 (kugou.com) — 单曲分享链接

    Args:
        url: 分享链接（支持短链和完整 URL）
        timeout: 请求超时秒数

    Returns:
        SongDetailInfo 包含歌曲详情和统计信息

    Raises:
        ValueError: URL 无法解析或平台不支持
        MusicClientError: 网络请求失败
    """
    platform = _detect_platform_from_url(url)
    try:
        if platform == "soda":
            return _scrape_soda_detail(url, timeout=timeout)
        elif platform == "netease":
            return _scrape_netease_detail(url, timeout=timeout)
        elif platform == "qq":
            return _scrape_qq_detail(url, timeout=timeout)
        elif platform == "kugou":
            return _scrape_kugou_song_detail(url, timeout=timeout)
        else:
            raise ValueError(
                f"不支持的平台 URL: {url}\n"
                "支持: 汽水音乐 (qishui.douyin.com)、"
                "网易云 (music.163.com)、QQ音乐 (y.qq.com)、"
                "酷狗音乐 (kugou.com)"
            )
    except requests.ConnectionError as e:
        raise MusicClientError(f"网络连接失败: {e}")
    except requests.Timeout:
        raise MusicClientError(
            f"请求超时 ({timeout}s)，可能遭遇限速，请稍后重试或用 --timeout 加长等待"
        )
    except requests.HTTPError as e:
        raise MusicClientError(f"HTTP 错误: {e}")


def get_playlist_detail_from_url(url: str, timeout: int = 15) -> PlaylistDetailInfo:
    """从歌单分享链接抓取歌单详情（数据监控，不含歌曲下载）。

    支持平台:
      - 汽水音乐 (qishui.douyin.com) — 含完整曲目统计数据

    Args:
        url: 歌单分享链接（支持短链和完整 URL）
        timeout: 请求超时秒数

    Returns:
        PlaylistDetailInfo 包含歌单详情和曲目列表

    Raises:
        ValueError: URL 无法解析或平台不支持
        MusicClientError: 网络请求失败
    """
    platform = _detect_platform_from_url(url)
    try:
        if platform == "soda":
            return _scrape_soda_playlist_detail(url, timeout=timeout)
        elif platform == "netease":
            return _scrape_netease_playlist_detail(url, timeout=timeout)
        elif platform == "qq":
            return _scrape_qq_playlist_detail(url, timeout=timeout)
        elif platform == "kugou":
            return _scrape_kugou_playlist_detail(url, timeout=timeout)
        else:
            raise ValueError(
                f"歌单详情暂不支持该平台 URL: {url}\n"
                "目前支持: 汽水音乐 (qishui.douyin.com)、"
                "网易云音乐 (music.163.com)、QQ音乐 (y.qq.com)、"
                "酷狗音乐 (kugou.com)"
            )
    except requests.ConnectionError as e:
        raise MusicClientError(f"网络连接失败: {e}")
    except requests.Timeout:
        raise MusicClientError(
            f"请求超时 ({timeout}s)，可能遭遇限速，请稍后重试或用 --timeout 加长等待"
        )
    except requests.HTTPError as e:
        raise MusicClientError(f"HTTP 错误: {e}")


# ─── Formatting Helpers ──────────────────────────────────────────────────────

def _source_emoji(source: str) -> str:
    """平台对应的 emoji 标识。"""
    emoji_map = {
        "netease": "🟣",
        "qq": "🟢",
        "kugou": "🔵",
        "kuwo": "🟠",
        "migu": "🔴",
        "bilibili": "🩷",
        "fivesing": "⭐",
        "soda": "🟡",
        "jamendo": "🎼",
        "joox": "🟤",
        "qianqian": "🔶",
    }
    return emoji_map.get(source, "🎵")


def _setup_chat_interactive(list_only: bool = False) -> None:
    """
    交互式选择飞书群聊并保存配置。

    Args:
        list_only: 仅列出群聊，不保存配置
    """
    # 检查环境变量
    app_id = os.environ.get("FEISHU_APP_ID", "")
    app_secret = os.environ.get("FEISHU_APP_SECRET", "")

    if not app_id or not app_secret:
        print("❌ 错误: 请先设置飞书应用凭证", file=sys.stderr)
        print("\n请执行以下命令设置环境变量:", file=sys.stderr)
        print('  export FEISHU_APP_ID="cli_xxx"', file=sys.stderr)
        print('  export FEISHU_APP_SECRET="xxx"', file=sys.stderr)
        print("\n然后重新运行: python3 music_toolkit.py setup-chat", file=sys.stderr)
        sys.exit(1)

    # 动态导入 feishu-toolkit
    try:
        sys.path.insert(0, str(Path(__file__).parent.parent / "feishu-toolkit"))
        from feishu_toolkit import FeishuClient
    except ImportError:
        print("❌ 错误: 无法导入 feishu-toolkit", file=sys.stderr)
        print("\n请确保 feishu-toolkit 在同级目录:", file=sys.stderr)
        print("  git clone https://github.com/mix9581/feishu-toolkit.git ../feishu-toolkit", file=sys.stderr)
        sys.exit(1)

    # 获取群聊列表
    print("🔍 正在获取群聊列表...")
    try:
        client = FeishuClient(app_id=app_id, app_secret=app_secret)
        chats = client.list_chats(page_size=100)
    except Exception as e:
        print(f"❌ 获取群聊列表失败: {e}", file=sys.stderr)
        sys.exit(1)

    if not chats:
        print("❌ 未找到任何群聊", file=sys.stderr)
        print("\n请确保:", file=sys.stderr)
        print("  1. 飞书应用已创建并获得权限", file=sys.stderr)
        print("  2. 机器人已被添加到至少一个群聊中", file=sys.stderr)
        sys.exit(1)

    # 显示群聊列表
    print(f"\n📋 找到 {len(chats)} 个群聊:\n")
    print(f"{'序号':<6} {'群聊名称':<40} {'Chat ID'}")
    print("─" * 80)

    for idx, chat in enumerate(chats, 1):
        chat_id = chat.get("chat_id", "")
        name = chat.get("name", "(未命名群聊)")
        # 截断过长的名称
        if len(name) > 38:
            name = name[:35] + "..."
        print(f"{idx:<6} {name:<40} {chat_id}")

    # 如果只是列出，直接返回
    if list_only:
        print("\n💡 提示: 使用以下命令保存配置:")
        print("   python3 music_toolkit.py setup-chat")
        return

    # 交互式选择
    print("\n" + "─" * 80)
    while True:
        try:
            choice = input(f"\n请选择群聊序号 (1-{len(chats)}, 或按 Ctrl+C 取消): ").strip()

            if not choice:
                continue

            idx = int(choice)
            if 1 <= idx <= len(chats):
                selected_chat = chats[idx - 1]
                break
            else:
                print(f"❌ 请输入 1 到 {len(chats)} 之间的数字")
        except ValueError:
            print("❌ 请输入有效的数字")
        except KeyboardInterrupt:
            print("\n\n⏹ 已取消")
            sys.exit(0)

    # 保存配置
    chat_id = selected_chat.get("chat_id", "")
    chat_name = selected_chat.get("name", "(未命名群聊)")

    print(f"\n✅ 已选择: {chat_name}")
    print(f"   Chat ID: {chat_id}")

    # 保存到配置文件
    config_dir = Path.home() / ".config" / "music-toolkit"
    config_dir.mkdir(parents=True, exist_ok=True)
    config_file = config_dir / "config.json"

    config = {}
    if config_file.exists():
        try:
            with open(config_file, "r", encoding="utf-8") as f:
                config = json.load(f)
        except Exception:
            pass

    config["feishu_default_chat_id"] = chat_id
    config["feishu_default_chat_name"] = chat_name

    try:
        with open(config_file, "w", encoding="utf-8") as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
        print(f"\n💾 配置已保存到: {config_file}")
    except Exception as e:
        print(f"\n⚠️  保存配置文件失败: {e}", file=sys.stderr)

    # 提示用户设置环境变量
    print("\n📝 请将以下内容添加到你的 shell 配置文件 (~/.bashrc 或 ~/.zshrc):")
    print(f'   export FEISHU_DEFAULT_CHAT_ID="{chat_id}"')
    print("\n或者每次使用时通过 --chat-id 参数指定:")
    print(f'   python3 music_toolkit.py download-send <song_id> <source> --chat-id "{chat_id}"')
    print("\n✨ 配置完成！现在可以使用飞书推送功能了。")


def _format_lyrics_preview(lyrics: str, max_lines: int = 4) -> str:
    """从 LRC 歌词中提取纯文本预览。"""
    lines = []
    for line in lyrics.splitlines():
        # Remove LRC time tags like [00:30.42]
        text = re.sub(r'\[\d{2}:\d{2}[\.\d]*\]', '', line).strip()
        # Skip metadata lines like [ti:...], [ar:...], [al:...], [by:...], [offset:...]
        if text and not re.match(r'^\[.+\]$', text):
            lines.append(text)
        if len(lines) >= max_lines:
            break
    return "\n".join(lines)


def _split_lyrics_for_doc(lyrics: str, chunk_size: int = 20) -> list[str]:
    """将 LRC 歌词分块用于飞书文档。"""
    lines = []
    for line in lyrics.splitlines():
        text = re.sub(r'\[\d{2}:\d{2}[\.\d]*\]', '', line).strip()
        if text and not re.match(r'^\[.+\]$', text):
            lines.append(text)

    chunks = []
    for i in range(0, len(lines), chunk_size):
        chunk = "\n".join(lines[i:i + chunk_size])
        if chunk:
            chunks.append(chunk)
    return chunks


def _format_song_table(songs: list[Song], show_index: bool = True) -> str:
    """格式化歌曲列表为终端友好的表格。"""
    if not songs:
        return "  (无结果)"

    lines = []
    for i, song in enumerate(songs, 1):
        prefix = f"  {i:>2}. " if show_index else "  "
        source_tag = f"[{song.source}]"
        line = f"{prefix}{song.name} - {song.artist}  ({song.duration_str})  {source_tag}"
        if song.size:
            line += f"  {song.size}"
        if song.bitrate and song.bitrate != "-":
            line += f"  {song.bitrate}"
        lines.append(line)
    return "\n".join(lines)


def _format_playlist_table(playlists: list[Playlist]) -> str:
    """格式化歌单列表为终端友好的表格。"""
    if not playlists:
        return "  (无结果)"

    lines = []
    for i, pl in enumerate(playlists, 1):
        line = f"  {i:>2}. {pl.name}  ({pl.track_count} 首)  by {pl.creator}  [{pl.source}]"
        lines.append(line)
    return "\n".join(lines)


# ─── CLI ──────────────────────────────────────────────────────────────────────

def _print_json(data):
    """打印 JSON 数据"""
    print(json.dumps(data, ensure_ascii=False, indent=2))


def _print_playlist_detail(playlist: "PlaylistDetailInfo", show_lyrics: bool = False) -> None:
    """格式化打印 PlaylistDetailInfo 的全部字段。"""
    sep = "─" * 60

    print(f"\n{sep}")
    print(f"📋  {playlist.title}")
    print(f"{sep}")
    print(f"   创建者:  {playlist.creator}")
    print(f"   歌曲数:  {playlist.track_count}")
    print(f"   平台:    {playlist.source_name} [{playlist.platform}]")
    print(f"   歌单ID:  {playlist.playlist_id}")
    if playlist.create_time:
        print(f"   创建:    {playlist.create_time}")
    if playlist.update_time:
        print(f"   更新:    {playlist.update_time}")
    if playlist.description:
        desc = playlist.description[:100] + ("..." if len(playlist.description) > 100 else "")
        print(f"   简介:    {desc}")
    if playlist.extra.get("play_count"):
        print(f"   播放:    {playlist.extra['play_count']:,}")
    if playlist.extra.get("collect_count"):
        print(f"   收藏:    {playlist.extra['collect_count']:,}")
    if playlist.extra.get("share_count"):
        print(f"   分享:    {playlist.extra['share_count']:,}")
    if playlist.cover:
        print(f"   封面:    {playlist.cover}")

    if playlist.tracks:
        has_album = any(t.album for t in playlist.tracks)
        has_dates = any(t.publish_date for t in playlist.tracks)
        has_favorites = any(t.favorites for t in playlist.tracks)
        has_comments = any(t.comments for t in playlist.tracks)
        has_shares = any(t.shares for t in playlist.tracks)
        has_links = any(t.resolved_url for t in playlist.tracks)

        # 动态表头
        header = f"  {'#':<4} {'歌名':<22} {'歌手':<12} {'时长':<6}"
        separ  = f"  {'─'*4} {'─'*22} {'─'*12} {'─'*6}"
        if has_album:
            header += f" {'专辑':<14}"
            separ  += f" {'─'*14}"
        if has_dates:
            header += f" {'发布日期':<10}"
            separ  += f" {'─'*10}"
        if has_favorites:
            header += f" {'收藏':>8}"
            separ  += f" {'─'*8}"
        if has_comments:
            header += f" {'评论':>7}"
            separ  += f" {'─'*7}"
        if has_shares:
            header += f" {'分享':>7}"
            separ  += f" {'─'*7}"

        print(f"\n{'─'*60}")
        print(header)
        print(separ)
        for i, t in enumerate(playlist.tracks, 1):
            row = f"  {i:<4} {t.name[:20]:<22} {t.artist[:10]:<12} {t.duration_str:<6}"
            if has_album:
                row += f" {t.album[:12]:<14}"
            if has_dates:
                row += f" {t.publish_date or '':<10}"
            if has_favorites:
                row += f" {t.favorites:>8,}"
            if has_comments:
                row += f" {t.comments:>7,}"
            if has_shares:
                row += f" {t.shares:>7,}"
            print(row)
            if has_links and t.resolved_url:
                print(f"       {t.resolved_url}")
            if show_lyrics and t.lyrics_lrc:
                for line in t.lyrics_text.splitlines()[:2]:
                    print(f"       ♪ {line}")

    print()


def _print_song_detail(detail: "SongDetailInfo", show_lyrics: bool = False) -> None:
    """格式化打印 SongDetailInfo 的全部字段。"""
    platform_name = PLATFORMS.get(detail.platform, detail.platform)
    sep = "─" * 52

    print(f"\n{sep}")
    print(f"🎵  {detail.name}")
    print(f"{sep}")

    # 基础信息
    print(f"   歌手:    {detail.artist}")
    if detail.album:
        print(f"   专辑:    {detail.album}" + (f"  (ID: {detail.album_id})" if detail.album_id else ""))
    print(f"   时长:    {detail.duration_str}  ({detail.duration}s)")
    print(f"   平台:    {platform_name} [{detail.platform}]")
    print(f"   歌曲ID:  {detail.song_id}")
    if detail.publish_date:
        print(f"   发布:    {detail.publish_date}")
    if detail.genre:
        print(f"   曲风:    {detail.genre}")
    if detail.language:
        print(f"   语言:    {detail.language}")
    if detail.composers:
        print(f"   作曲:    {detail.composers}")
    if detail.lyricists:
        print(f"   作词:    {detail.lyricists}")
    if detail.qualities:
        print(f"   音质:    {detail.qualities}")

    # 封面
    if detail.cover:
        print(f"   封面:    {detail.cover}")

    # 音频直链
    if detail.audio_url:
        url_display = detail.audio_url[:80] + "..." if len(detail.audio_url) > 80 else detail.audio_url
        print(f"   音频:    {url_display}")
        print(f"            (CDN 直链，临时有效约 24h)")

    # 互动统计
    print(f"\n📊 互动统计:")
    print(f"   收藏:    {detail.favorites:,}")
    print(f"   评论:    {detail.comments:,}")
    print(f"   分享:    {detail.shares:,}")
    if detail.plays:
        print(f"   播放:    {detail.plays:,}")

    # 链接
    if detail.resolved_url and detail.resolved_url != detail.share_url:
        print(f"\n🔗 解析链接:")
        print(f"   {detail.resolved_url}")

    # 歌词
    if detail.lyrics_lrc:
        if show_lyrics:
            print(f"\n📝 歌词 (LRC):\n")
            for line in detail.lyrics_lrc.splitlines():
                print(f"   {line}")
        else:
            lines = detail.lyrics_text.splitlines()
            preview = "\n".join(f"   {l}" for l in lines[:4])
            print(f"\n📝 歌词预览:\n{preview}")
            if len(lines) > 4:
                print(f"   ... 共 {len(lines)} 行  (--lyrics 查看完整歌词)")

    print(f"\n💡 下载: python music_toolkit.py parse-url {detail.share_url} --download")




def main():
    parser = argparse.ArgumentParser(
        prog="music_toolkit",
        description="🎵 Music Toolkit — 音乐搜索/下载/飞书推送",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  %(prog)s search "晴天"                     # 全平台搜索
  %(prog)s search "晴天" --source qq          # 指定平台搜索
  %(prog)s search-playlist "周杰伦精选"       # 搜索歌单
  %(prog)s detail 0042rlGx2WHBrG qq          # 歌曲详情
  %(prog)s lyrics 0042rlGx2WHBrG qq          # 获取歌词
  %(prog)s download 0042rlGx2WHBrG qq        # 下载歌曲
  %(prog)s music-detail "https://qishui.douyin.com/s/CODE/"  # 分享链接详情
  %(prog)s playlist-detail "https://qishui.douyin.com/s/CODE/"  # 歌单详情（数据监控）
  %(prog)s playlist-detail "..." --lyrics     # 同时下载歌词
  %(prog)s switch-source --name 晴天 --artist 周杰伦  # 换源
  %(prog)s platforms                          # 列出平台
  %(prog)s push-song 0042rlGx2WHBrG qq       # 推送飞书
  %(prog)s push-search "晴天"                 # 搜索+推送

环境变量:
  GO_MUSIC_DL_URL         go-music-dl 地址 (默认 http://localhost:8080)
  DOWNLOAD_DIR            下载目录 (默认 ./downloads)
  FEISHU_APP_ID           飞书 App ID (推送用)
  FEISHU_APP_SECRET       飞书 App Secret (推送用)
  FEISHU_DEFAULT_CHAT_ID  默认飞书群 ID (推送用)
        """,
    )
    sub = parser.add_subparsers(dest="command")

    # ── search ─────────────────────────────────────────────────────────────
    p = sub.add_parser("search", help="搜索歌曲")
    p.add_argument("keyword", help="搜索关键词")
    p.add_argument("--source", action="append", dest="sources", help="指定平台 (可多次)")
    p.add_argument("--json", action="store_true", dest="as_json", help="输出 JSON")

    # ── search-playlist ────────────────────────────────────────────────────
    p = sub.add_parser("search-playlist", help="搜索歌单")
    p.add_argument("keyword", help="搜索关键词")
    p.add_argument("--source", action="append", dest="sources", help="指定平台 (可多次)")
    p.add_argument("--json", action="store_true", dest="as_json", help="输出 JSON")

    # ── detail ─────────────────────────────────────────────────────────────
    p = sub.add_parser("detail", help="获取歌曲详情 (inspect + lyrics)")
    p.add_argument("song_id", help="歌曲 ID")
    p.add_argument("source", help="音源平台")
    p.add_argument("--json", action="store_true", dest="as_json", help="输出 JSON")

    # ── lyrics ─────────────────────────────────────────────────────────────
    p = sub.add_parser("lyrics", help="获取歌词")
    p.add_argument("song_id", help="歌曲 ID")
    p.add_argument("source", help="音源平台")

    # ── download ───────────────────────────────────────────────────────────
    p = sub.add_parser("download", help="下载歌曲")
    p.add_argument("song_id", help="歌曲 ID")
    p.add_argument("source", help="音源平台")
    p.add_argument("--name", default="", help="歌曲名 (文件名用)")
    p.add_argument("--artist", default="", help="歌手名 (文件名用)")
    p.add_argument("--embed", action="store_true", help="嵌入封面")
    p.add_argument("--dir", dest="save_dir", help="保存目录")

    # ── switch-source ──────────────────────────────────────────────────────
    p = sub.add_parser("switch-source", help="换源搜索")
    p.add_argument("--name", required=True, help="歌曲名")
    p.add_argument("--artist", required=True, help="歌手名")
    p.add_argument("--source", default="", help="当前平台 (排除)")
    p.add_argument("--json", action="store_true", dest="as_json", help="输出 JSON")

    # ── playlist ───────────────────────────────────────────────────────────
    p = sub.add_parser("playlist", help="获取歌单歌曲列表")
    p.add_argument("playlist_id", help="歌单 ID")
    p.add_argument("source", help="音源平台")
    p.add_argument("--json", action="store_true", dest="as_json", help="输出 JSON")

    # ── download-playlist ─────────────────────────────────────────────────
    p = sub.add_parser("download-playlist", help="批量下载歌单所有歌曲 (自动换源)")
    p.add_argument("playlist_id", help="歌单 ID")
    p.add_argument("source", help="音源平台")
    p.add_argument("--dir", dest="save_dir", help="保存目录")
    p.add_argument("--webhook", help="下载完成后推送到飞书 webhook URL")
    p.add_argument("--send-chat", dest="send_chat_id", metavar="CHAT_ID",
                   help="下载完成后发送文件到飞书群 (打包zip，大文件自动用云盘分片上传)")
    p.add_argument("--lyrics-doc", action="store_true",
                   help="同时创建飞书歌词文档 (每首歌名为 H1 标题)")
    p.add_argument("--json", action="store_true", dest="as_json", help="输出 JSON")

    # ── download-send ─────────────────────────────────────────────────────
    p = sub.add_parser("download-send", help="下载单曲并发送到飞书群 (mp3 + txt 歌词)")
    p.add_argument("song_id", help="歌曲 ID")
    p.add_argument("source", help="音源平台")
    p.add_argument("--name", default="", help="歌曲名 (用于文件名)")
    p.add_argument("--artist", default="", help="歌手名 (用于文件名)")
    p.add_argument("--dir", dest="save_dir", help="保存目录")
    p.add_argument("--chat-id", dest="send_chat_id", metavar="CHAT_ID",
                   help="目标飞书群 ID (不指定则用 FEISHU_DEFAULT_CHAT_ID)")

    # ── music-detail ──────────────────────────────────────────────────────────
    p = sub.add_parser("music-detail", help="从分享链接抓取歌曲详情（收藏/评论/分享/封面等）")
    p.add_argument("url", help="歌曲分享链接（汽水/网易云/QQ音乐等）")
    p.add_argument("--json", action="store_true", dest="as_json", help="输出 JSON")
    p.add_argument("--lyrics", action="store_true", help="同时输出完整歌词")
    p.add_argument("--related", action="store_true", help="同时列出同艺人/相关曲目（仅汽水，无需额外请求）")
    p.add_argument("--timeout", type=int, default=15, help="请求超时秒数 (默认 15)")

    # ── music-detail-batch ────────────────────────────────────────────────
    p = sub.add_parser("music-detail-batch", help="批量抓取多条歌曲详情（支持多 URL 或歌单文件）")
    p.add_argument("urls", nargs="*", help="歌曲分享链接（可多个，空格分隔）")
    p.add_argument("--file", dest="url_file", metavar="FILE",
                   help="从文件读取 URL（每行一个）")
    p.add_argument("--delay", type=float, default=2.0,
                   help="请求间隔秒数，避免限速 (默认 2.0)")
    p.add_argument("--json", action="store_true", dest="as_json", help="输出 JSON 数组")
    p.add_argument("--timeout", type=int, default=15, help="单条请求超时秒数 (默认 15)")

    # ── parse-url ──────────────────────────────────────────────────────────
    p = sub.add_parser("parse-url", help="解析音乐分享链接 → 详情 + 歌词 + 下载")
    p.add_argument("url", help="音乐分享链接 (网易云/QQ音乐/酷狗 等)")
    p.add_argument("--download", action="store_true", help="同时下载歌曲文件")
    p.add_argument("--dir", dest="save_dir", help="下载保存目录")
    p.add_argument("--webhook", help="推送到飞书 webhook URL")
    p.add_argument("--json", action="store_true", dest="as_json", help="输出 JSON")

    # ── playlist-detail ────────────────────────────────────────────────────
    p = sub.add_parser("playlist-detail", help="从歌单分享链接抓取歌单详情（数据监控，不含歌曲下载）")
    p.add_argument("url", help="歌单分享链接（汽水音乐等）")
    p.add_argument("--json", action="store_true", dest="as_json", help="输出 JSON")
    p.add_argument("--lyrics", action="store_true",
                   help="同时下载每首歌的歌词文件 (.lrc)，保存到 --dir 目录")
    p.add_argument("--dir", dest="save_dir", default="./downloads",
                   help="歌词保存目录 (默认 ./downloads，--lyrics 时生效)")
    p.add_argument("--delay", type=float, default=1.5,
                   help="歌词请求间隔秒数 (默认 1.5)")
    p.add_argument("--timeout", type=int, default=15, help="请求超时秒数 (默认 15)")

    # ── platforms ──────────────────────────────────────────────────────────
    sub.add_parser("platforms", help="列出支持的平台")

    # ── push-song ──────────────────────────────────────────────────────────
    p = sub.add_parser("push-song", help="获取歌曲详情并推送飞书卡片")
    p.add_argument("song_id", help="歌曲 ID")
    p.add_argument("source", help="音源平台")
    p.add_argument("--chat-id", help="飞书群 ID")

    # ── push-search ────────────────────────────────────────────────────────
    p = sub.add_parser("push-search", help="搜索歌曲并推送结果到飞书")
    p.add_argument("keyword", help="搜索关键词")
    p.add_argument("--source", action="append", dest="sources", help="指定平台 (可多次)")
    p.add_argument("--chat-id", help="飞书群 ID")

    # ── push-playlist ──────────────────────────────────────────────────────
    p = sub.add_parser("push-playlist", help="获取歌单并推送到飞书")
    p.add_argument("playlist_id", help="歌单 ID")
    p.add_argument("source", help="音源平台")
    p.add_argument("--chat-id", help="飞书群 ID")

    # ── push-playlist-detail ───────────────────────────────────────────────
    p = sub.add_parser("push-playlist-detail", help="抓取歌单详情（数据监控）并推送到飞书群")
    p.add_argument("url", help="歌单分享链接（汽水音乐等）")
    p.add_argument("--chat-id", help="飞书群 ID（不指定则用 FEISHU_DEFAULT_CHAT_ID）")
    p.add_argument("--max-tracks", type=int, default=0, help="卡片最多显示曲目数（默认 0 = 全部显示）")
    p.add_argument(
        "--sort", dest="sort_by", default="",
        choices=["likes", "comments", "shares", "date"],
        help="排序字段: likes/comments/shares/date（默认歌单原序）",
    )
    p.add_argument(
        "--desc", dest="sort_desc", action="store_true", default=True,
        help="降序排列（默认）",
    )
    p.add_argument("--asc", dest="sort_desc", action="store_false", help="升序排列")
    p.add_argument("--timeout", type=int, default=15, help="请求超时秒数（默认 15）")
    p.add_argument("--with-doc", action="store_true", dest="with_doc",
                   help="推送后生成 CSV 文件发送到群（包含所有曲目完整数据）")

    # ── playlist-to-table ─────────────────────────────────────────────────
    p = sub.add_parser(
        "playlist-to-table",
        help="抓取歌单数据并创建飞书多维表格（汽水/网易云/QQ音乐通用）",
    )
    p.add_argument("url", help="歌单分享链接")
    p.add_argument("--chat-id", help="可选：创建完成后向该飞书群发送表格链接卡片")
    p.add_argument(
        "--sort", dest="sort_by", default="",
        choices=["likes", "comments", "shares", "date"],
        help="排序字段: likes/comments/shares/date（默认歌单原序）",
    )
    p.add_argument("--desc", dest="sort_desc", action="store_true", default=True,
                   help="降序排列（默认）")
    p.add_argument("--asc", dest="sort_desc", action="store_false", help="升序排列")
    p.add_argument("--timeout", type=int, default=15, help="请求超时秒数（默认 15）")

    # ── push-webhook ───────────────────────────────────────────────────────
    p = sub.add_parser("push-webhook", help="搜索歌曲并推送到飞书 webhook")
    p.add_argument("keyword", help="搜索关键词")
    p.add_argument("webhook_url", help="飞书 webhook URL")
    p.add_argument("--source", action="append", dest="sources", help="指定平台 (可多次)")

    # ── send-to-chat ──────────────────────────────────────────────────────
    p = sub.add_parser("send-to-chat", help="发送歌曲文件到飞书群 (单首直接发，多首打包zip)")
    p.add_argument("path", help="文件路径或目录路径 (目录会发送所有音频文件)")
    p.add_argument("--chat-id", help="飞书群 chat_id")
    p.add_argument("--zip-name", default="", help="压缩包名称 (多文件时)")
    p.add_argument("--include-lrc", action="store_true", help="同时包含歌词文件 (.lrc)")

    # ── setup-chat ────────────────────────────────────────────────────────
    p = sub.add_parser("setup-chat", help="交互式选择飞书群聊 (自动保存配置)")
    p.add_argument("--list-only", action="store_true", help="仅列出群聊，不保存配置")

    # ── cookie ────────────────────────────────────────────────────────────
    p = sub.add_parser("cookie", help="管理平台 Cookie (用于访问会员歌曲/完整歌单)")
    cookie_sub = p.add_subparsers(dest="cookie_action", help="Cookie 操作")

    # cookie list
    p_list = cookie_sub.add_parser("list", help="查看当前所有 Cookie")
    p_list.add_argument("--json", action="store_true", dest="as_json", help="输出 JSON")

    # cookie set
    p_set = cookie_sub.add_parser("set", help="设置平台 Cookie")
    p_set.add_argument("source", help="平台代码 (netease/qq/kugou/kuwo/migu/bilibili)")
    p_set.add_argument("cookie", help="Cookie 字符串")

    # cookie delete
    p_del = cookie_sub.add_parser("delete", help="删除平台 Cookie")
    p_del.add_argument("source", nargs="+", help="平台代码 (可多个)")

    # cookie clear
    cookie_sub.add_parser("clear", help="清除所有 Cookie")

    # cookie load
    p_load = cookie_sub.add_parser("load", help="从文件或环境变量加载 Cookie")
    p_load.add_argument("--file", dest="cookie_file", help="JSON 文件路径 (默认 .music_cookies.json)")
    p_load.add_argument("--env", action="store_true", help="从环境变量加载")

    # cookie save
    p_save = cookie_sub.add_parser("save", help="保存当前 Cookie 到文件")
    p_save.add_argument("--file", dest="cookie_file", help="JSON 文件路径 (默认 .music_cookies.json)")

    args = parser.parse_args()
    if not args.command:
        parser.print_help()
        return

    try:
        _run_command(args)
    except MusicClientError as e:
        print(f"❌ 错误: {e}", file=sys.stderr)
        sys.exit(1)
    except KeyboardInterrupt:
        print("\n⏹ 已取消", file=sys.stderr)
        sys.exit(130)


def _run_command(args):
    """执行 CLI 子命令。"""
    client = MusicClient()

    if args.command == "search":
        songs = client.search_songs(args.keyword, sources=args.sources)
        if args.as_json:
            _print_json([s.to_dict() for s in songs])
        else:
            print(f"\n🔍 搜索: {args.keyword} (共 {len(songs)} 首)\n")
            print(_format_song_table(songs))
            if songs:
                print(f"\n💡 获取详情: python music_toolkit.py detail {songs[0].id} {songs[0].source}")

    elif args.command == "search-playlist":
        playlists = client.search_playlists(args.keyword, sources=args.sources)
        if args.as_json:
            _print_json([p.to_dict() for p in playlists])
        else:
            print(f"\n📋 歌单搜索: {args.keyword} (共 {len(playlists)} 个)\n")
            print(_format_playlist_table(playlists))
            if playlists:
                print(f"\n💡 查看歌单: python music_toolkit.py playlist {playlists[0].id} {playlists[0].source}")

    elif args.command == "detail":
        # First search to get basic info (name/artist)
        inspect_result = client.inspect(args.song_id, args.source)
        lyrics = ""
        try:
            lyrics = client.get_lyrics(args.song_id, args.source)
        except Exception:
            pass

        if args.as_json:
            _print_json({
                "id": args.song_id,
                "source": args.source,
                "inspect": inspect_result.to_dict(),
                "lyrics": lyrics,
            })
        else:
            print(f"\n🎵 歌曲详情: {args.song_id} [{args.source}]\n")
            print(f"  有效: {'✅ 是' if inspect_result.valid else '❌ 否'}")
            print(f"  大小: {inspect_result.size}")
            print(f"  码率: {inspect_result.bitrate}")
            if inspect_result.url:
                # Truncate URL for display
                url_display = inspect_result.url[:80] + "..." if len(inspect_result.url) > 80 else inspect_result.url
                print(f"  URL:  {url_display}")
            if lyrics:
                preview = _format_lyrics_preview(lyrics, max_lines=6)
                print(f"\n📝 歌词预览:\n{preview}")

    elif args.command == "lyrics":
        lyrics = client.get_lyrics(args.song_id, args.source)
        print(lyrics)

    elif args.command == "download":
        print(f"⏬ 正在下载: {args.song_id} [{args.source}] ...")
        filepath = client.download(
            args.song_id, args.source,
            name=args.name, artist=args.artist,
            embed=args.embed, save_dir=args.save_dir,
        )
        print(f"✅ 已保存: {filepath}")

    elif args.command == "switch-source":
        song = client.switch_source(args.name, args.artist, source=args.source)
        if song:
            if args.as_json:
                _print_json(song.to_dict())
            else:
                print(f"\n🔄 换源结果:\n")
                print(f"  歌名: {song.name}")
                print(f"  歌手: {song.artist}")
                print(f"  平台: {song.source_name} [{song.source}]")
                print(f"  时长: {song.duration_str}")
                if song.album:
                    print(f"  专辑: {song.album}")
                if song.link:
                    print(f"  链接: {song.link}")
                if song.score > 0:
                    print(f"  匹配度: {song.score:.0%}")
                print(f"\n💡 下载: python music_toolkit.py download {song.id} {song.source}")
        else:
            print("❌ 未找到匹配的换源结果")

    elif args.command == "playlist":
        songs = client.get_playlist_songs(args.playlist_id, args.source)
        if args.as_json:
            _print_json([s.to_dict() for s in songs])
        else:
            print(f"\n📋 歌单: {args.playlist_id} [{args.source}] (共 {len(songs)} 首)\n")
            print(_format_song_table(songs))

    elif args.command == "download-playlist":
        print(f"\n📋 获取歌单: {args.playlist_id} [{args.source}] ...")
        songs = client.get_playlist_songs(args.playlist_id, args.source)
        if not songs:
            print("❌ 歌单为空或无法获取")
            return

        print(f"   共 {len(songs)} 首，开始批量下载...\n")

        def _on_progress(idx, total, name, status):
            print(f"  [{idx:>2}/{total}] {name} — {status}")

        results = client.download_playlist(
            args.playlist_id, args.source,
            save_dir=args.save_dir,
            on_progress=_on_progress,
        )

        # 统计
        success = [r for r in results if r.success]
        failed = [r for r in results if not r.success]
        switched = [r for r in success if r.actual_source != args.source]

        print(f"\n{'─' * 50}")
        print(f"📊 下载完成: {len(success)}/{len(results)} 成功")
        if switched:
            print(f"   🔄 换源下载: {len(switched)} 首")
            for r in switched:
                src_name = PLATFORMS.get(r.actual_source, r.actual_source)
                print(f"      {r.song.name} → {src_name}")
        if failed:
            print(f"   ❌ 失败: {len(failed)} 首")
            for r in failed:
                print(f"      {r.song.name}: {r.error}")

        if args.as_json:
            _print_json([r.to_dict() for r in results])

        # 推送到飞书 webhook
        if hasattr(args, "webhook") and args.webhook and success:
            print(f"\n📨 正在推送下载报告到飞书...")
            _push_playlist_download_report(
                args.webhook, args.playlist_id, args.source,
                results, songs,
            )

        # 发送文件到飞书群
        if hasattr(args, "send_chat_id") and args.send_chat_id and success:
            file_paths = []
            for r in success:
                if r.filepath:
                    file_paths.append(r.filepath)
                if r.lrc_path:
                    file_paths.append(r.lrc_path)
                if r.txt_path:
                    file_paths.append(r.txt_path)
            if file_paths:
                print(f"\n📤 正在发送 {len(file_paths)} 个文件到飞书群...")
                try:
                    pusher = FeishuPusher()
                    pusher.send_song_files(
                        file_paths,
                        chat_id=args.send_chat_id,
                        zip_name=f"歌单_{args.playlist_id}",
                    )
                    print(f"✅ 文件已发送到飞书群")
                except Exception as e:
                    print(f"❌ 发送失败: {e}", file=sys.stderr)

        # 创建飞书歌词文档
        if hasattr(args, "lyrics_doc") and args.lyrics_doc and success:
            print(f"\n📝 正在创建歌词文档...")
            try:
                # 为成功的歌曲获取歌词
                enriched_songs = []
                for r in success:
                    try:
                        enriched = client.enrich_song(r.song)
                        enriched_songs.append(enriched)
                    except Exception:
                        enriched_songs.append(r.song)

                pusher = FeishuPusher()
                doc_url = pusher.create_playlist_lyrics_doc(
                    enriched_songs,
                    title=f"歌单歌词 - {args.playlist_id}",
                )
                print(f"✅ 歌词文档: {doc_url}")

                # 如果有 send_chat_id，也发送文档链接到群
                if hasattr(args, "send_chat_id") and args.send_chat_id:
                    card = pusher._client.build_card(
                        "📝 歌单歌词文档",
                        [
                            pusher._client.card_markdown(
                                f"共 {len(enriched_songs)} 首歌词已整理\n"
                                f"每首歌名为 H1 标题，方便搜索查找"
                            ),
                            pusher._client.card_button("📄 打开文档", doc_url),
                        ],
                        color="purple",
                    )
                    pusher._client.send_card(args.send_chat_id, card)
                    print(f"✅ 歌词文档链接已推送到飞书群")
            except Exception as e:
                print(f"❌ 创建歌词文档失败: {e}", file=sys.stderr)

    elif args.command == "download-send":
        # 构造 Song 对象
        song = Song(
            id=args.song_id, source=args.source,
            name=args.name or args.song_id,
            artist=args.artist or "",
            duration=0, cover="",
        )

        # 如果没有提供 name/artist，尝试通过搜索获取
        if not args.name:
            try:
                results = client.search_songs(args.song_id, sources=[args.source])
                if results:
                    song = results[0]
            except Exception:
                pass

        print(f"\n🎵 下载: {song.name} - {song.artist} [{song.source}]")
        result = client._download_single_with_fallback(song, save_dir=args.save_dir)

        if not result.success:
            print(f"❌ 下载失败: {result.error}", file=sys.stderr)
            return

        print(f"✅ 下载完成: {result.filepath}")
        if result.txt_path:
            print(f"   歌词: {result.txt_path}")

        # 发送到飞书群
        try:
            pusher = FeishuPusher()
            cid = args.send_chat_id or ""
            if result.filepath:
                print(f"\n📤 发送 mp3 到飞书群...")
                pusher.send_song_files([result.filepath], chat_id=cid)
                print(f"✅ mp3 已发送")
            if result.txt_path:
                print(f"📤 发送歌词到飞书群...")
                pusher.send_song_files([result.txt_path], chat_id=cid)
                print(f"✅ 歌词已发送")
        except Exception as e:
            print(f"❌ 飞书发送失败: {e}", file=sys.stderr)

    elif args.command == "music-detail":
        detail = get_song_detail_from_url(args.url, timeout=args.timeout)
        if args.as_json:
            _print_json(detail.to_dict())
        else:
            _print_song_detail(detail, show_lyrics=args.lyrics)
            if args.related and detail.platform == "soda":
                print(f"\n{'─' * 52}")
                print("🎵 同艺人 / 相关曲目 (来自页面嵌入数据，无额外请求)\n")
                related = get_soda_related_tracks(args.url, timeout=args.timeout)
                for i, r in enumerate(related, 1):
                    print(
                        f"  {i:>2}. {r.name} — {r.artist}"
                        f"  ({r.duration_str})  "
                        f"收藏:{r.favorites:,}  评论:{r.comments:,}  分享:{r.shares:,}"
                    )
                    print(f"       ID: {r.song_id}  专辑: {r.album}  发布: {r.publish_date}")

    elif args.command == "music-detail-batch":
        import time as _time

        # Collect URLs
        urls: list[str] = list(args.urls or [])
        if args.url_file:
            try:
                with open(args.url_file, encoding="utf-8") as uf:
                    for line in uf:
                        line = line.strip()
                        if line and not line.startswith("#"):
                            urls.append(line)
            except FileNotFoundError:
                print(f"❌ 文件不存在: {args.url_file}", file=sys.stderr)
                return

        if not urls:
            print("❌ 请提供至少一个 URL 或 --file 参数", file=sys.stderr)
            return

        print(f"\n📋 批量抓取 {len(urls)} 首歌曲详情\n")
        results = []
        for idx, url in enumerate(urls, 1):
            prefix = f"[{idx}/{len(urls)}]"
            print(f"{prefix} 抓取: {url[:60]}...")
            try:
                d = get_song_detail_from_url(url, timeout=args.timeout)
                results.append(d)
                print(
                    f"  ✅ {d.name} — {d.artist}"
                    f"  收藏:{d.favorites:,}  评论:{d.comments:,}  分享:{d.shares:,}"
                )
            except Exception as e:
                print(f"  ❌ 失败: {e}")
                results.append(None)

            if idx < len(urls):
                _time.sleep(args.delay)

        if args.as_json:
            _print_json([r.to_dict() if r else None for r in results])
        else:
            valid = [r for r in results if r]
            print(f"\n{'─' * 52}")
            print(f"📊 完成: {len(valid)}/{len(urls)} 成功\n")
            if valid:
                print(f"  {'#':<4} {'歌名':<22} {'歌手':<12} {'时长':<6} {'收藏':>8} {'评论':>7} {'分享':>7}")
                print(f"  {'─'*4} {'─'*22} {'─'*12} {'─'*6} {'─'*8} {'─'*7} {'─'*7}")
                for i, r in enumerate(valid, 1):
                    print(
                        f"  {i:<4} {r.name[:20]:<22} {r.artist[:10]:<12} {r.duration_str:<6}"
                        f" {r.favorites:>8,} {r.comments:>7,} {r.shares:>7,}"
                    )

    elif args.command == "parse-url":
        # 把 URL 作为搜索关键词，go-music-dl 会自动解析
        songs = client.search_songs(args.url)
        if not songs:
            print(f"❌ 无法解析链接: {args.url}")
            return

        song = songs[0]
        enriched = client.enrich_song(song)

        if args.as_json:
            _print_json(enriched.to_dict())
        else:
            print(f"\n🎵 {enriched.name}")
            print(f"   歌手:   {enriched.artist}")
            if enriched.album:
                print(f"   专辑:   {enriched.album}")
            print(f"   时长:   {enriched.duration_str}")
            print(f"   平台:   {enriched.source_name} [{enriched.source}]")
            print(f"   ID:     {enriched.id}")
            if enriched.size:
                print(f"   大小:   {enriched.size}")
            if enriched.bitrate and enriched.bitrate != "-":
                print(f"   音质:   {enriched.bitrate}")

            if enriched.lyrics:
                print(f"\n📝 歌词:\n")
                # 打印完整歌词（去掉时间标签）
                for line in enriched.lyrics.splitlines():
                    text = re.sub(r'\[\d{2}:\d{2}[\.\d]*\]', '', line).strip()
                    if text and not re.match(r'^\[.+\]$', text):
                        print(f"   {text}")
            else:
                print("\n📝 暂无歌词")

        # 下载
        if args.download:
            # 先检查当前平台资源是否可用
            inspect_result = client.inspect(enriched.id, enriched.source, duration=enriched.duration)
            download_song = enriched

            if not inspect_result.valid:
                print(f"\n⚠️  {enriched.source_name} 资源不可用，尝试换源...")
                alt = client.switch_source(
                    enriched.name, enriched.artist,
                    source=enriched.source, duration=enriched.duration,
                )
                if alt and alt.id:
                    alt_inspect = client.inspect(alt.id, alt.source, duration=alt.duration)
                    if alt_inspect.valid:
                        download_song = Song(
                            id=alt.id, source=alt.source,
                            name=alt.name or enriched.name,
                            artist=alt.artist or enriched.artist,
                            duration=alt.duration, cover=alt.cover,
                            album=alt.album, extra=alt.extra,
                            url=alt_inspect.url, size=alt_inspect.size,
                            bitrate=alt_inspect.bitrate,
                        )
                        print(f"   找到可用源: {download_song.source_name} [{download_song.source}]"
                              f"  {download_song.size}")
                    else:
                        print(f"❌ 换源后仍无法下载，该歌曲可能有版权限制")
                        return
                else:
                    print(f"❌ 未找到可用的替代源，该歌曲可能有版权限制")
                    return

            print(f"\n⏬ 正在下载...")
            filepath = client.download(
                download_song.id, download_song.source,
                name=download_song.name, artist=download_song.artist,
                cover=download_song.cover, extra=download_song.extra,
                save_dir=args.save_dir,
            )
            print(f"✅ 已保存: {filepath}")

            # 同时下载歌词文件
            if enriched.lyrics:
                lrc_path = client.download_lyrics_file(
                    enriched.id, enriched.source,
                    name=enriched.name, artist=enriched.artist,
                    save_dir=args.save_dir,
                )
                print(f"✅ 歌词: {lrc_path}")

        # 推送飞书 webhook
        if hasattr(args, "webhook") and args.webhook:
            result = push_to_webhook(args.webhook, enriched)
            print(f"\n📨 已推送到飞书 (卡片 + 歌词)")

    elif args.command == "playlist-detail":
        import time as _time

        playlist = get_playlist_detail_from_url(args.url, timeout=args.timeout)

        if args.as_json:
            _print_json(playlist.to_dict())
        else:
            _print_playlist_detail(playlist)

        # 下载歌词（每首单独请求，尽力而为）
        if args.lyrics and playlist.tracks:
            print(f"\n📝 正在下载 {len(playlist.tracks)} 首歌曲歌词 ...")
            save_dir = Path(args.save_dir)
            save_dir.mkdir(parents=True, exist_ok=True)
            ok = 0
            for idx, track in enumerate(playlist.tracks, 1):
                if not track.song_id:
                    continue
                try:
                    # Construct a share URL to fetch LRC from the go-music-dl backend
                    client = MusicClient()
                    lrc_content = client.get_lyrics(track.song_id, "soda")
                    if lrc_content:
                        filename = _sanitize_filename(
                            _build_filename(track.name, track.artist, "lrc")
                        )
                        lrc_path = save_dir / filename
                        lrc_path.write_text(lrc_content, encoding="utf-8")
                        _save_txt_lyrics(lrc_path)
                        ok += 1
                        print(f"  [{idx:>2}/{len(playlist.tracks)}] ✅ {track.name}")
                    else:
                        print(f"  [{idx:>2}/{len(playlist.tracks)}] ⚠️  {track.name} (无歌词)")
                except Exception as e:
                    print(f"  [{idx:>2}/{len(playlist.tracks)}] ❌ {track.name}: {e}")
                if idx < len(playlist.tracks):
                    _time.sleep(args.delay)
            print(f"\n📊 歌词下载完成: {ok}/{len(playlist.tracks)} 首  → {save_dir}/")

    elif args.command == "platforms":
        print("\n🎵 支持的平台:\n")
        for code, name in PLATFORMS.items():
            print(f"  {code:<12} {name}")
        print(f"\n  共 {len(PLATFORMS)} 个平台")

    elif args.command == "push-song":
        # Enrich with inspect + lyrics
        inspect_result = client.inspect(args.song_id, args.source)
        lyrics = ""
        try:
            lyrics = client.get_lyrics(args.song_id, args.source)
        except Exception:
            pass

        song = Song(
            id=args.song_id,
            source=args.source,
            name=args.song_id,  # Will use ID as name if not found
            artist="",
            duration=0,
            cover="",
            url=inspect_result.url,
            size=inspect_result.size,
            bitrate=inspect_result.bitrate,
            lyrics=lyrics,
        )

        pusher = FeishuPusher()
        pusher.push_song_card(song, chat_id=args.chat_id)
        print("✅ 歌曲详情已推送到飞书")

    elif args.command == "push-search":
        songs = client.search_songs(args.keyword, sources=args.sources)
        if not songs:
            print(f"❌ 未找到: {args.keyword}")
            return
        # Enrich first few songs with inspect
        enriched = []
        for song in songs[:5]:
            try:
                enriched.append(client.enrich_song(song))
            except Exception:
                enriched.append(song)
        # Keep the rest as-is
        enriched.extend(songs[5:])

        pusher = FeishuPusher()
        pusher.push_search_results(enriched, args.keyword, chat_id=args.chat_id)
        print(f"✅ 搜索结果 ({len(songs)} 首) 已推送到飞书")

    elif args.command == "push-playlist":
        songs = client.get_playlist_songs(args.playlist_id, args.source)
        playlist = Playlist(
            id=args.playlist_id,
            source=args.source,
            name=f"歌单 {args.playlist_id}",
            cover="",
            track_count=len(songs),
        )
        pusher = FeishuPusher()
        pusher.push_playlist_card(playlist, songs=songs, chat_id=args.chat_id)
        print(f"✅ 歌单 ({len(songs)} 首) 已推送到飞书")

    elif args.command == "push-playlist-detail":
        print(f"📋 抓取歌单数据: {args.url[:60]}...")
        playlist = get_playlist_detail_from_url(args.url, timeout=args.timeout)
        print(f"   {playlist.title} — {len(playlist.tracks)} 首")

        pusher = FeishuPusher()
        pusher.push_playlist_detail_card(
            playlist,
            chat_id=args.chat_id,
            max_tracks=args.max_tracks,
            sort_by=args.sort_by,
            sort_desc=args.sort_desc,
            with_doc=getattr(args, "with_doc", False),
        )
        sort_hint = f"（排序: {args.sort_by} {'降序' if args.sort_desc else '升序'}）" if args.sort_by else ""
        print(f"✅ 歌单卡片已推送到飞书群{sort_hint}")

    elif args.command == "playlist-to-table":
        print(f"📋 抓取歌单数据: {args.url[:60]}...")
        playlist = get_playlist_detail_from_url(args.url, timeout=args.timeout)
        print(f"   {playlist.title} — {len(playlist.tracks)} 首 [{playlist.source_name}]")

        pusher = FeishuPusher()
        sort_hint = f"（{args.sort_by} {'降序' if args.sort_desc else '升序'}）" if args.sort_by else ""
        print(f"   正在创建飞书多维表格{sort_hint}...")
        bitable_url = pusher.create_playlist_bitable(
            playlist,
            sort_by=args.sort_by,
            sort_desc=args.sort_desc,
        )
        print(f"✅ 多维表格已创建")
        print(f"   {bitable_url}")

        # 可选：向群发送表格链接卡片
        chat_id = args.chat_id or os.environ.get("FEISHU_DEFAULT_CHAT_ID", "")
        if chat_id:
            c = pusher._client
            cid = pusher._resolve_chat_id(args.chat_id)
            sort_label = {"likes": "点赞", "comments": "评论",
                          "shares": "分享", "date": "日期"}.get(args.sort_by, "")
            sort_note = (
                f"\n排序: {sort_label} {'↓' if args.sort_desc else '↑'}"
                if sort_label else ""
            )
            card = c.build_card(
                f"{playlist.title} — 数据表格",
                [
                    c.card_markdown(
                        f"**{playlist.title}**  ·  {playlist.creator}  ·  "
                        f"{len(playlist.tracks)} 首 [{playlist.source_name}]{sort_note}"
                    ),
                    c.card_button("打开多维表格", bitable_url),
                ],
                color="wathet",
            )
            c.send_card(cid, card)
            print(f"✅ 表格链接已发送到飞书群")

    elif args.command == "push-webhook":
        songs = client.search_songs(args.keyword, sources=args.sources)
        if not songs:
            print(f"❌ 未找到: {args.keyword}")
            return

        # 找到最匹配的歌曲（优先原唱，非 Cover 版本）
        song = songs[0]
        for s in songs:
            if "Cover" not in s.name and "cover" not in s.name.lower():
                song = s
                break

        # 获取完整详情
        enriched = client.enrich_song(song)

        # 推送到 webhook
        result = push_to_webhook(args.webhook_url, enriched)
        print(f"✅ 已推送到飞书 webhook")
        print(f"   歌曲: {enriched.name}")
        print(f"   歌手: {enriched.artist}")
        print(f"   平台: {enriched.source_name}")
        print(f"   响应: {result.get('msg', 'success')}")

    elif args.command == "setup-chat":
        _setup_chat_interactive(list_only=args.list_only)

    elif args.command == "cookie":
        if not hasattr(args, "cookie_action") or not args.cookie_action:
            print("❌ 请指定 cookie 操作: list/set/delete/clear/load/save", file=sys.stderr)
            print("   示例: python music_toolkit.py cookie list")
            sys.exit(1)

        if args.cookie_action == "list":
            cookies = client.get_cookies()
            if args.as_json:
                _print_json(cookies)
            else:
                if not cookies:
                    print("\n🍪 当前没有设置任何 Cookie\n")
                    print("💡 设置 Cookie:")
                    print('   python music_toolkit.py cookie set netease "MUSIC_U=xxx"')
                    print('   python music_toolkit.py cookie load --env')
                    print('   python music_toolkit.py cookie load --file .music_cookies.json')
                else:
                    print("\n🍪 当前 Cookie 配置:\n")
                    for source, cookie in cookies.items():
                        platform_name = PLATFORMS.get(source, source)
                        cookie_preview = cookie[:50] + "..." if len(cookie) > 50 else cookie
                        print(f"  {source:<12} ({platform_name})")
                        print(f"               {cookie_preview}")
                    print(f"\n  共 {len(cookies)} 个平台")

        elif args.cookie_action == "set":
            client.set_cookie(args.source, args.cookie)
            platform_name = PLATFORMS.get(args.source, args.source)
            print(f"✅ 已设置 {platform_name} ({args.source}) 的 Cookie")

        elif args.cookie_action == "delete":
            client.clear_cookies(args.source)
            for src in args.source:
                platform_name = PLATFORMS.get(src, src)
                print(f"✅ 已删除 {platform_name} ({src}) 的 Cookie")

        elif args.cookie_action == "clear":
            client.clear_cookies()
            print("✅ 已清除所有 Cookie")

        elif args.cookie_action == "load":
            if args.env:
                loaded = client.load_cookies_from_env()
                if loaded:
                    print(f"✅ 从环境变量加载了 {len(loaded)} 个 Cookie:")
                    for src in loaded.keys():
                        platform_name = PLATFORMS.get(src, src)
                        print(f"   - {platform_name} ({src})")
                else:
                    print("⚠️  未找到任何 Cookie 环境变量")
                    print("\n💡 环境变量命名规则:")
                    for src, env_var in COOKIE_ENV_VARS.items():
                        print(f"   export {env_var}=\"your_cookie_here\"")
            else:
                filepath = args.cookie_file or DEFAULT_COOKIE_FILE
                loaded = client.load_cookies_from_file(filepath)
                if loaded:
                    print(f"✅ 从文件加载了 {len(loaded)} 个 Cookie: {filepath}")
                    for src in loaded.keys():
                        platform_name = PLATFORMS.get(src, src)
                        print(f"   - {platform_name} ({src})")
                else:
                    print(f"⚠️  文件不存在或为空: {filepath}")
                    print(f"\n💡 创建 JSON 文件:")
                    print(f'   echo \'{{"netease": "MUSIC_U=xxx", "qq": "uin=xxx"}}\' > {filepath}')

        elif args.cookie_action == "save":
            filepath = args.cookie_file or DEFAULT_COOKIE_FILE
            client.save_cookies_to_file(filepath)
            cookies = client.get_cookies()
            print(f"✅ 已保存 {len(cookies)} 个 Cookie 到: {filepath}")

    elif args.command == "send-to-chat":
        target = Path(args.path)
        audio_exts = {".mp3", ".m4a", ".flac", ".wav", ".ogg", ".opus", ".aac", ".wma"}
        lrc_ext = {".lrc"}

        if target.is_file():
            file_paths = [target]
        elif target.is_dir():
            # 收集目录下的音频文件
            file_paths = sorted([
                f for f in target.iterdir()
                if f.suffix.lower() in audio_exts
            ])
            if args.include_lrc:
                lrc_paths = sorted([
                    f for f in target.iterdir()
                    if f.suffix.lower() in lrc_ext
                ])
                file_paths.extend(lrc_paths)
        else:
            print(f"❌ 路径不存在: {args.path}", file=sys.stderr)
            sys.exit(1)

        if not file_paths:
            print(f"❌ 未找到音频文件: {args.path}", file=sys.stderr)
            sys.exit(1)

        print(f"📤 准备发送 {len(file_paths)} 个文件到飞书群...")
        for f in file_paths:
            size_mb = f.stat().st_size / (1024 * 1024)
            print(f"   {f.name}  ({size_mb:.1f} MB)")

        if len(file_paths) > 1:
            zip_name = args.zip_name or target.name or "songs"
            print(f"\n📦 多文件，打包为: {zip_name}.zip")

        pusher = FeishuPusher()
        result = pusher.send_song_files(
            file_paths,
            chat_id=args.chat_id,
            zip_name=args.zip_name or (target.name if target.is_dir() else ""),
        )
        print(f"✅ 文件已发送到飞书群")


if __name__ == "__main__":
    main()
